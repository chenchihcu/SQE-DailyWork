"""Shared read-only Excel workbook helpers for master-data import previews."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from zipfile import BadZipFile


def normalize_excel_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_excel_header(value: Any) -> str:
    return normalize_excel_cell(value).lower()


def find_excel_column_index(headers: list[Any], accepted_names: set[str]) -> int | None:
    accepted = {name.lower() for name in accepted_names}
    for index, header in enumerate(headers):
        if normalize_excel_header(header) in accepted:
            return index
    return None


def read_excel_workbook_rows(
    file_path: str | Path,
    *,
    error_type: type[Exception],
) -> tuple[list[Any], list[tuple[int, list[Any]]]]:
    """Read the first worksheet header row and non-empty data rows."""
    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise error_type("僅支援 .xlsx 檔案。")
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise error_type(f"無法開啟 Excel 檔案：{exc}") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, BadZipFile, ValueError, InvalidFileException) as exc:
        raise error_type(f"無法開啟 Excel 檔案：{exc}") from exc

    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None or not any(normalize_excel_cell(cell) for cell in header_row):
            return [], []
        data_rows: list[tuple[int, list[Any]]] = []
        for row_number, row_values in enumerate(rows_iter, start=2):
            values = list(row_values)
            if any(normalize_excel_cell(cell) for cell in values):
                data_rows.append((row_number, values))
        return list(header_row), data_rows
    finally:
        workbook.close()
