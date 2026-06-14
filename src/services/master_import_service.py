from __future__ import annotations

import sqlite3
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from database.product_stage import (
    PRODUCT_STAGE_MASS_PRODUCTION,
    normalize_product_stage_ui,
)


STATUS_CAN_ADD = "可新增"
STATUS_CAN_UPDATE = "可更新"
STATUS_SKIPPED = "已一致略過"
STATUS_BLOCKED = "錯誤擋下"
IMPORT_TYPE_PRODUCT_MASTER = "product_master"
IMPORT_BATCH_COMPLETED = "completed"
IMPORT_BATCH_BLOCKED = "blocked"
IMPORT_BATCH_SKIPPED = "skipped"

PRODUCT_CODE_HEADERS = {"料號", "產品料號", "product_code", "item_no"}
PRODUCT_NAME_HEADERS = {"品名", "產品名稱", "product_name"}
SUPPLIER_NAME_HEADERS = {"供應商", "主供應商", "supplier_name"}
PRODUCT_STAGE_HEADERS = {"階段", "產品階段", "product_stage"}


class MasterImportError(ValueError):
    """Raised when a master-data import source cannot be parsed safely."""


@dataclass(frozen=True)
class ProductMasterImportRow:
    row_number: int
    status: str
    product_code: str
    product_name: str
    supplier_name: str
    product_stage: str
    current_product_name: str
    current_supplier_name: str
    message: str
    will_create_supplier: bool = False
    will_assign_supplier: bool = False

    @property
    def can_add(self) -> bool:
        return self.status == STATUS_CAN_ADD

    @property
    def can_update(self) -> bool:
        return self.status == STATUS_CAN_UPDATE

    @property
    def is_error(self) -> bool:
        return self.status == STATUS_BLOCKED


@dataclass(frozen=True)
class ProductMasterImportPreview:
    rows: list[ProductMasterImportRow]
    file_errors: list[str]

    @property
    def add_count(self) -> int:
        return sum(1 for row in self.rows if row.can_add)

    @property
    def update_count(self) -> int:
        return sum(1 for row in self.rows if row.can_update)

    @property
    def supplier_create_count(self) -> int:
        names = {
            row.supplier_name
            for row in self.rows
            if row.will_create_supplier and row.supplier_name
        }
        return len(names)

    @property
    def skipped_count(self) -> int:
        return sum(1 for row in self.rows if row.status == STATUS_SKIPPED)

    @property
    def error_count(self) -> int:
        return len(self.file_errors) + sum(1 for row in self.rows if row.is_error)

    @property
    def can_import(self) -> bool:
        return self.error_count == 0

    @property
    def has_writes(self) -> bool:
        return self.add_count > 0 or self.update_count > 0 or self.supplier_create_count > 0

    def rows_to_write(self) -> list[ProductMasterImportRow]:
        return [row for row in self.rows if row.can_add or row.can_update]


@dataclass(frozen=True)
class ProductMasterImportResult:
    added_count: int
    updated_count: int
    supplier_created_count: int
    skipped_count: int
    backup_path: Path | None
    batch_id: str | None = None


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return _normalize_cell(value).lower()


def _find_column_index(headers: list[Any], accepted_names: set[str]) -> int | None:
    accepted = {name.lower() for name in accepted_names}
    for index, header in enumerate(headers):
        if _normalize_header(header) in accepted:
            return index
    return None


def _read_workbook_rows(file_path: str | Path) -> tuple[list[Any], list[tuple[int, list[Any]]]]:
    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise MasterImportError("僅支援 .xlsx 檔案。")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, BadZipFile, InvalidFileException) as exc:
        raise MasterImportError(f"無法開啟 Excel 檔案：{exc}") from exc
    except ValueError as exc:
        raise MasterImportError(f"Excel 檔案格式無法讀取：{exc}") from exc

    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None or not any(_normalize_cell(cell) for cell in header_row):
            return [], []
        data_rows: list[tuple[int, list[Any]]] = []
        for row_number, row_values in enumerate(rows_iter, start=2):
            values = list(row_values)
            if any(_normalize_cell(cell) for cell in values):
                data_rows.append((row_number, values))
        return list(header_row), data_rows
    finally:
        workbook.close()


def _cell(values: list[Any], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return _normalize_cell(values[index])


def _supplier_by_name(conn: sqlite3.Connection, supplier_name: str) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, supplier_name, is_active
        FROM suppliers
        WHERE supplier_name = ?
        LIMIT 1
        """,
        (supplier_name,),
    ).fetchone()


def _product_by_code(
    conn: sqlite3.Connection, product_code: str, supplier_id: str | None = None
) -> sqlite3.Row | None:
    if supplier_id:
        return conn.execute(
            """
            SELECT
                p.id,
                p.product_code,
                p.product_name,
                p.product_stage,
                p.supplier_id,
                s.supplier_name
            FROM products p
            LEFT JOIN suppliers s ON s.id = p.supplier_id
            WHERE p.product_code = ? AND p.supplier_id = ? AND p.is_active = 1
            LIMIT 1
            """,
            (product_code, supplier_id),
        ).fetchone()
    return conn.execute(
        """
        SELECT
            p.id,
            p.product_code,
            p.product_name,
            p.product_stage,
            p.supplier_id,
            s.supplier_name
        FROM products p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        WHERE p.product_code = ? AND p.is_active = 1
        LIMIT 1
        """,
        (product_code,),
    ).fetchone()


def preview_product_master_import(
    conn: sqlite3.Connection, file_path: str | Path
) -> ProductMasterImportPreview:
    headers, data_rows = _read_workbook_rows(file_path)
    file_errors: list[str] = []
    preview_rows: list[ProductMasterImportRow] = []
    if not headers:
        return ProductMasterImportPreview(rows=[], file_errors=["Excel 第一列需包含欄位標題。"])

    code_index = _find_column_index(headers, PRODUCT_CODE_HEADERS)
    name_index = _find_column_index(headers, PRODUCT_NAME_HEADERS)
    supplier_index = _find_column_index(headers, SUPPLIER_NAME_HEADERS)
    stage_index = _find_column_index(headers, PRODUCT_STAGE_HEADERS)
    if code_index is None:
        file_errors.append("找不到料號欄位，請使用「料號」、「產品料號」或 product_code。")
    if name_index is None:
        file_errors.append("找不到品名欄位，請使用「品名」、「產品名稱」或 product_name。")
    if supplier_index is None:
        file_errors.append("找不到主供應商欄位，請使用「供應商」、「主供應商」或 supplier_name。")
    if file_errors:
        return ProductMasterImportPreview(rows=[], file_errors=file_errors)
    if not data_rows:
        return ProductMasterImportPreview(rows=[], file_errors=["Excel 沒有可匯入的資料列。"])

    normalized_rows: list[tuple[int, str, str, str, str]] = []
    code_counts: dict[str, int] = {}
    for row_number, values in data_rows:
        product_code = _cell(values, code_index)
        product_name = _cell(values, name_index)
        supplier_name = _cell(values, supplier_index)
        product_stage = normalize_product_stage_ui(_cell(values, stage_index))
        if not product_stage:
            product_stage = PRODUCT_STAGE_MASS_PRODUCTION
        normalized_rows.append(
            (row_number, product_code, product_name, supplier_name, product_stage)
        )
        if product_code:
            code_counts[product_code] = code_counts.get(product_code, 0) + 1

    for row_number, product_code, product_name, supplier_name, product_stage in normalized_rows:
        if not product_code:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    "",
                    product_name,
                    supplier_name,
                    product_stage,
                    "",
                    "",
                    "料號不可空白。",
                )
            )
            continue
        if not product_name:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    product_code,
                    "",
                    supplier_name,
                    product_stage,
                    "",
                    "",
                    "品名不可空白。",
                )
            )
            continue
        if not supplier_name:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    product_code,
                    product_name,
                    "",
                    product_stage,
                    "",
                    "",
                    "主供應商不可空白；ERP 匯入需先指定供應商。",
                )
            )
            continue
        if code_counts.get(product_code, 0) > 1:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    product_code,
                    product_name,
                    supplier_name,
                    product_stage,
                    "",
                    "",
                    "Excel 內部有重複料號，請先整理來源資料。",
                )
            )
            continue

        supplier = _supplier_by_name(conn, supplier_name)
        if supplier is not None and not bool(supplier["is_active"]):
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    product_code,
                    product_name,
                    supplier_name,
                    product_stage,
                    "",
                    supplier_name,
                    "主供應商已停用，請先在基礎資料啟用或更正 ERP 匯出資料。",
                )
            )
            continue

        supplier_id_for_lookup = str(supplier["id"]) if supplier is not None else None
        existing = _product_by_code(conn, product_code, supplier_id=supplier_id_for_lookup)
        will_create_supplier = supplier is None
        if existing is None:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_CAN_ADD,
                    product_code,
                    product_name,
                    supplier_name,
                    product_stage,
                    "",
                    "",
                    "將新增至共用產品主檔。",
                    will_create_supplier=will_create_supplier,
                )
            )
            continue

        current_supplier = str(existing["supplier_name"] or "").strip()
        current_name = str(existing["product_name"] or "").strip()
        if current_supplier and current_supplier != supplier_name:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    product_code,
                    product_name,
                    supplier_name,
                    product_stage,
                    current_name,
                    current_supplier,
                    "料號已屬於不同主供應商，請人工確認後再調整。",
                )
            )
            continue

        needs_name_update = current_name != product_name
        needs_supplier_assign = not current_supplier
        if needs_name_update or needs_supplier_assign:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_CAN_UPDATE,
                    product_code,
                    product_name,
                    supplier_name,
                    product_stage,
                    current_name,
                    current_supplier,
                    "將更新品名或補齊主供應商；不改寫事件或倉庫不合格品資料。",
                    will_create_supplier=will_create_supplier,
                    will_assign_supplier=needs_supplier_assign,
                )
            )
        else:
            preview_rows.append(
                ProductMasterImportRow(
                    row_number,
                    STATUS_SKIPPED,
                    product_code,
                    product_name,
                    supplier_name,
                    product_stage,
                    current_name,
                    current_supplier,
                    "共用產品主檔已一致，匯入時略過。",
                )
            )
    return ProductMasterImportPreview(rows=preview_rows, file_errors=file_errors)


def _database_file_path(conn: sqlite3.Connection) -> Path | None:
    for row in conn.execute("PRAGMA database_list").fetchall():
        name = row["name"] if isinstance(row, sqlite3.Row) else row[1]
        file_name = row["file"] if isinstance(row, sqlite3.Row) else row[2]
        if name != "main" or not file_name or file_name == ":memory:":
            continue
        path = Path(str(file_name))
        if path.exists():
            return path
    return None


def _create_database_backup(conn: sqlite3.Connection) -> Path | None:
    source_path = _database_file_path(conn)
    if source_path is None:
        return None
    backup_dir = source_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"sqe-before-product-master-import-{timestamp}.db"
    suffix = 1
    while backup_path.exists():
        backup_path = backup_dir / f"sqe-before-product-master-import-{timestamp}-{suffix}.db"
        suffix += 1
    backup_conn = sqlite3.connect(backup_path)
    try:
        conn.backup(backup_conn)
    finally:
        backup_conn.close()
    return backup_path


def _source_file_parts(source_file: str | Path | None) -> tuple[str, str]:
    if source_file is None:
        return "", ""
    path = Path(source_file)
    return path.name, str(path)


def _insert_import_batch(
    conn: sqlite3.Connection,
    preview: ProductMasterImportPreview,
    *,
    source_file: str | Path | None,
    status: str,
    added_count: int = 0,
    updated_count: int = 0,
    supplier_created_count: int = 0,
    skipped_count: int | None = None,
    error_count: int | None = None,
    backup_path: Path | None = None,
    message: str = "",
    started_at: str | None = None,
    completed_at: str | None = None,
) -> str:
    batch_id = uuid.uuid4().hex
    source_name, source_path = _source_file_parts(source_file)
    created_at = completed_at or datetime.now().isoformat(timespec="seconds")
    started = started_at or created_at
    final_skipped_count = preview.skipped_count if skipped_count is None else skipped_count
    final_error_count = preview.error_count if error_count is None else error_count
    conn.execute(
        """
        INSERT INTO import_batches(
            id, import_type, source_file, source_path, status, total_rows,
            added_count, updated_count, supplier_created_count, skipped_count,
            error_count, backup_path, message, started_at, completed_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            batch_id,
            IMPORT_TYPE_PRODUCT_MASTER,
            source_name,
            source_path,
            status,
            len(preview.rows),
            added_count,
            updated_count,
            supplier_created_count,
            final_skipped_count,
            final_error_count,
            str(backup_path or ""),
            message,
            started,
            created_at,
        ),
    )
    for row in preview.rows:
        conn.execute(
            """
            INSERT INTO import_batch_rows(
                id, batch_id, row_number, status, product_code, product_name,
                supplier_name, product_stage, message, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                uuid.uuid4().hex,
                batch_id,
                row.row_number,
                row.status,
                row.product_code,
                row.product_name,
                row.supplier_name,
                row.product_stage,
                row.message,
                created_at,
            ),
        )
    return batch_id


def record_product_master_import_rejection(
    conn: sqlite3.Connection,
    preview: ProductMasterImportPreview,
    *,
    source_file: str | Path | None = None,
) -> str:
    """Persist a rejected ERP/product-master preview without touching master data."""
    if preview.can_import:
        raise ValueError("只有預覽失敗的匯入批次可登錄為 blocked。")
    message = "預覽發現錯誤，未寫入 suppliers/products。"
    conn.execute("BEGIN")
    try:
        batch_id = _insert_import_batch(
            conn,
            preview,
            source_file=source_file,
            status=IMPORT_BATCH_BLOCKED,
            error_count=preview.error_count,
            message=message,
        )
        conn.commit()
        return batch_id
    except sqlite3.Error:
        conn.rollback()
        raise


def _ensure_supplier(conn: sqlite3.Connection, supplier_name: str, now: str) -> tuple[str, bool]:
    supplier = _supplier_by_name(conn, supplier_name)
    if supplier is not None:
        return str(supplier["id"]), False
    supplier_id = uuid.uuid4().hex
    conn.execute(
        """
        INSERT INTO suppliers(
            id, supplier_name, contact_name, department, phone, contact_email,
            is_active, created_at, updated_at
        ) VALUES (?, ?, '', '', '', '', 1, ?, ?)
        """,
        (supplier_id, supplier_name, now, now),
    )
    return supplier_id, True


def apply_product_master_import(
    conn: sqlite3.Connection,
    preview: ProductMasterImportPreview,
    *,
    source_file: str | Path | None = None,
) -> ProductMasterImportResult:
    if not preview.can_import:
        raise ValueError("匯入清單仍有錯誤，請修正 Excel 後重新預覽。")
    if not preview.has_writes:
        batch_id = _insert_import_batch(
            conn,
            preview,
            source_file=source_file,
            status=IMPORT_BATCH_SKIPPED,
            skipped_count=preview.skipped_count,
            error_count=0,
            message="共用產品與供應商主檔已一致，未寫入 master data。",
        )
        conn.commit()
        return ProductMasterImportResult(
            added_count=0,
            updated_count=0,
            supplier_created_count=0,
            skipped_count=preview.skipped_count,
            backup_path=None,
            batch_id=batch_id,
        )

    now = datetime.now().isoformat(timespec="seconds")
    backup_path = _create_database_backup(conn)
    supplier_created_count = 0
    added_count = 0
    updated_count = 0
    conn.execute("BEGIN")
    try:
        for row in preview.rows_to_write():
            supplier_id, created_supplier = _ensure_supplier(conn, row.supplier_name, now)
            if created_supplier:
                supplier_created_count += 1

            if row.can_add:
                conn.execute(
                    """
                    INSERT INTO products(
                        id, product_code, product_name, product_stage, supplier_id,
                        secondary_supplier_id, is_active, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, NULL, 1, ?, ?)
                    """,
                    (
                        uuid.uuid4().hex,
                        row.product_code,
                        row.product_name,
                        row.product_stage or PRODUCT_STAGE_MASS_PRODUCTION,
                        supplier_id,
                        now,
                        now,
                    ),
                )
                added_count += 1
                continue

            existing = _product_by_code(conn, row.product_code, supplier_id=supplier_id)
            if existing is None:
                raise sqlite3.IntegrityError(f"Product disappeared: {row.product_code}")
            conn.execute(
                """
                UPDATE products
                SET product_name = ?,
                    supplier_id = COALESCE(supplier_id, ?),
                    updated_at = ?
                WHERE id = ?
                """,
                (row.product_name, supplier_id, now, existing["id"]),
            )
            updated_count += 1
        batch_id = _insert_import_batch(
            conn,
            preview,
            source_file=source_file,
            status=IMPORT_BATCH_COMPLETED,
            added_count=added_count,
            updated_count=updated_count,
            supplier_created_count=supplier_created_count,
            skipped_count=preview.skipped_count,
            error_count=0,
            backup_path=backup_path,
            message="匯入完成；只寫入 suppliers/products 共用主檔。",
            started_at=now,
        )
        conn.commit()
        return ProductMasterImportResult(
            added_count=added_count,
            updated_count=updated_count,
            supplier_created_count=supplier_created_count,
            skipped_count=preview.skipped_count,
            backup_path=backup_path,
            batch_id=batch_id,
        )
    except sqlite3.Error:
        conn.rollback()
        raise
