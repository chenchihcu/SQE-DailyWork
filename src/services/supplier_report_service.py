"""Supplier report export with explicit source-separated worksheets."""

from __future__ import annotations

from database import connection as _connection
from database import repository
from database.repo_helpers import format_current_action_text
from services import supplier_360_service


def export_supplier_report(
    file_path: str,
    supplier_id: str,
    start_date: str,
    end_date: str,
) -> tuple[bool, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        summary = supplier_360_service.get_supplier_summary(supplier_id)
        supplier = summary.get("supplier") or {}
        workbook = Workbook()
        overview = workbook.active
        overview.title = "供應商摘要"
        scorecard = supplier_360_service.get_supplier_scorecard(
            supplier_id, start_date, end_date
        )
        rows = [
            ("供應商", supplier.get("supplier_name", "")),
            ("報告區間", f"{start_date} ～ {end_date}"),
            ("目前評級", scorecard.get("grade", "")),
            ("未結異常", summary.get("open_anomaly_count", 0)),
            ("逾期異常", summary.get("overdue_anomaly_count", 0)),
            ("重複警示", summary.get("repeat_flagged_anomaly_count", 0)),
            ("近 90 日 NCR", summary.get("ncr_90d_count", 0)),
            ("最近訪廠", summary.get("latest_visit_date", "")),
        ]
        for row in rows:
            overview.append(row)
        overview["A1"].font = Font(bold=True)

        anomaly_rows = _enriched_supplier_anomalies(supplier_id)
        _append_table(
            workbook,
            "異常統計",
            anomaly_rows,
            (
                "anomaly_no",
                "anomaly_date",
                "problem_desc",
                "status",
                "responsible_person",
                "due_date",
                "overdue",
                "current_action_text",
                "root_cause_status",
                "corrective_action_status",
                "verification_result",
            ),
        )
        _append_table(
            workbook,
            "訪廠紀錄",
            supplier_360_service.list_supplier_visits(supplier_id),
            ("visit_date", "summary", "visitor_name", "status", "work_order_no"),
        )
        _append_table(
            workbook,
            "不合格品統計",
            supplier_360_service.list_supplier_defects(supplier_id),
            ("defect_no", "event_date", "item_no", "product_name", "defect_desc", "status"),
        )
        score_sheet = workbook.create_sheet("評分摘要")
        for key, value in scorecard.items():
            score_sheet.append((key, value))
        workbook.save(file_path)
        return True, f"已輸出供應商報告：{file_path}"
    except Exception as exc:
        return False, f"供應商報告輸出失敗：{exc}"


def _enriched_supplier_anomalies(supplier_id: str) -> list[dict]:
    base_rows = supplier_360_service.list_supplier_anomalies(supplier_id)
    enriched: list[dict] = []
    with _connection.get_connection() as conn:
        for row in base_rows:
            anomaly_id = conn.execute(
                "SELECT id FROM anomalies WHERE anomaly_no = ? LIMIT 1",
                (str(row.get("anomaly_no") or ""),),
            ).fetchone()
            if anomaly_id is None:
                enriched.append(dict(row))
                continue
            aid = str(anomaly_id["id"])
            overview = repository.get_anomaly_overview_card(conn, aid)
            enriched.append(
                {
                    **row,
                    "overdue": "是" if overview.get("overdue") else "否",
                    "current_action_text": format_current_action_text(
                        overview.get("current_action"),
                        include_owner_due=True,
                    ),
                    "root_cause_status": overview.get("root_cause_status") or "尚未開始",
                    "corrective_action_status": overview.get(
                        "corrective_action_status"
                    ) or "—",
                    "verification_result": overview.get("verification_result") or "—",
                }
            )
    return enriched


def _append_table(workbook, title: str, rows: list[dict], keys: tuple[str, ...]) -> None:
    from openpyxl.styles import Font

    sheet = workbook.create_sheet(title)
    sheet.append(keys)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(key, "") for key in keys])
