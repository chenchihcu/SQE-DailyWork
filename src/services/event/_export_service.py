"""Event PDF/Excel export and report generation."""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING

logger = logging.getLogger(__name__)

from services import event_pdf_exporter

from . import _anomaly_service
from ._anomaly_markdown import OVERVIEW_FIELDS
from ._helpers import _month_now
from . import _query_service

_OVERVIEW_LABELS = dict(OVERVIEW_FIELDS)

if TYPE_CHECKING:
    from PySide6.QtGui import QImage


def _event_pdf_payload(row: dict) -> dict:
    event_id = str(row.get("event_id") or "").strip()
    if not event_id:
        raise ValueError("Event id is required")
    event_type = str(row.get("event_type") or "").strip().upper()
    if event_type != "ANOMALY":
        raise ValueError("Event type is required")
    return _anomaly_service.get_anomaly_detail(event_id)


def default_event_pdf_filename(row: dict) -> str:
    detail = _event_pdf_payload(row)
    return event_pdf_exporter.default_event_pdf_filename(row, detail)


def _run_event_pdf_export(
    path: str,
    row: dict,
    delegate,
    *,
    failure_log: str,
    failure_prefix: str,
) -> tuple[bool, str]:
    """Shared payload-resolve → delegate → exception-to-(False, msg) wrapper
    for the full/brief PDF export entry points (audit finding D17)."""
    try:
        detail = _event_pdf_payload(row)
        return delegate(path, row, detail)
    except Exception as exc:
        logger.exception(failure_log)
        return False, f"{failure_prefix}{exc}"


def export_event_pdf(path: str, row: dict) -> tuple[bool, str]:
    return _run_event_pdf_export(
        path,
        row,
        event_pdf_exporter.export_event_pdf,
        failure_log="PDF 匯出失敗",
        failure_prefix="匯出失敗：",
    )


def export_brief_event_pdf(path: str, row: dict) -> tuple[bool, str]:
    return _run_event_pdf_export(
        path,
        row,
        event_pdf_exporter.export_brief_event_pdf,
        failure_log="精簡版 PDF 匯出失敗",
        failure_prefix="匯出精簡版失敗：",
    )


def render_brief_event_image(row: dict) -> "QImage | None":
    """將精簡報告渲染為 QImage 供 LINE 剪貼簿圖片傳送。"""
    try:
        detail = _event_pdf_payload(row)
        return event_pdf_exporter.render_brief_event_to_image(
            row,
            detail,
        )
    except Exception:
        logger.exception("渲染精簡報告圖片失敗")
        return None


def _stats_period_label(yyyymm: str) -> str:
    period_key = str(yyyymm or "").strip().upper()
    today = date.today()
    if period_key == "ALL":
        return "全期項目"
    if period_key == "YEAR":
        return f"{today.year} 年度"
    if period_key == "HALF_YEAR":
        half_label = "上半年" if today.month <= 6 else "下半年"
        return f"{today.year} {half_label}"

    month = str(yyyymm or _month_now()).strip()
    if len(month) == 7 and "-" in month:
        month = month.replace("-", "")
    if len(month) == 6 and month.isdigit():
        return f"{month[:4]}-{month[4:]}"
    return month


def _append_hypothesis_export_sheet(
    workbook,
    anomaly_rows: list[dict],
    *,
    output_parent: Path,
    warnings: list[str],
) -> list[Path]:
    """Add a hypothesis overview sheet with optional embedded tree PNGs (max 12)."""
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font

    from services.event import _anomaly_workbench_service
    from services.event._hypothesis_tree_png import (
        HYPOTHESIS_EXCEL_PNG_LIMIT,
        format_hypothesis_tree_text,
        render_hypothesis_tree_png,
    )

    candidates = [
        row
        for row in anomaly_rows
        if int(row.get("hypothesis_count") or 0) > 0
    ]
    if not candidates:
        return []

    sheet = workbook.create_sheet("原因假設")
    sheet.views.sheetView[0].showGridLines = True
    headers = ["異常單號", "供應商", "假設數", "樹狀圖 / 文字摘要"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    png_budget = HYPOTHESIS_EXCEL_PNG_LIMIT
    skipped_png = 0
    temp_png_paths: list[Path] = []
    for row in candidates:
        anomaly_id = str(row.get("event_id") or "").strip()
        ref_no = str(row.get("ref_no") or "")
        supplier = str(row.get("supplier_name") or "")
        count = int(row.get("hypothesis_count") or 0)
        hypotheses = _anomaly_workbench_service.list_hypotheses(anomaly_id)
        text_summary = format_hypothesis_tree_text(hypotheses)
        row_idx = sheet.max_row + 1
        sheet.append([ref_no, supplier, count, text_summary[:2000]])
        if png_budget <= 0:
            skipped_png += 1
            continue
        png_path = output_parent / f"temp_hypothesis_{anomaly_id}.png"
        if render_hypothesis_tree_png(hypotheses, png_path):
            img = Image(str(png_path))
            img.width = 360
            img.height = 220
            anchor = f"D{row_idx}"
            sheet.add_image(img, anchor)
            png_budget -= 1
            temp_png_paths.append(png_path)
        else:
            warnings.append(f"原因假設樹 PNG 未產生：{ref_no}")

    if skipped_png:
        warnings.append(
            f"原因假設 PNG 僅嵌入前 {HYPOTHESIS_EXCEL_PNG_LIMIT} 案；"
            f"其餘 {skipped_png} 案僅輸出文字摘要"
        )
    return temp_png_paths


def export_monthly_excel(path: str, yyyymm: str) -> tuple[bool, str]:
    import pandas as pd
    try:
        month = yyyymm or _month_now()
        stats = _query_service.get_monthly_stats(month)
        rows = _query_service.list_events({"yyyymm": month})
        summary_df = pd.DataFrame(
            [
                {
                    "月份": _stats_period_label(month),
                    "本月異常數": stats["anomaly_count"],
                    "訪廠數": stats["visit_count"],
                    "結案數": stats["closed_anomaly_count"],
                    "未結案數": stats["open_anomaly_count"],
                    "結案率(%)": stats["close_rate_pct"],
                    "異常/訪廠比": stats["anomaly_visit_ratio"],
                    "供應商覆蓋數": stats["supplier_coverage_count"],
                }
            ]
        )
        ranking_df = pd.DataFrame(
            [
                {
                    "排名": idx,
                    "供應商": row["supplier_name"],
                    "異常數": row["anomaly_count"],
                    "訪廠數": row["visit_count"],
                    "結案數": row["closed_anomaly_count"],
                    "未結案數": row["open_anomaly_count"],
                    "結案率(%)": row["close_rate_pct"],
                }
                for idx, row in enumerate(stats["top_suppliers_by_anomaly"], start=1)
            ],
            columns=["排名", "供應商", "異常數", "訪廠數", "結案數", "未結案數", "結案率(%)"],
        )
        resp_stats = _query_service.get_responsible_person_stats(month)
        resp_ranking_df = pd.DataFrame(
            [
                {
                    "排名": idx,
                    "責任人": row["responsible_person"],
                    "異常數": row.get("anomaly_count", row.get("total_count", 0)),
                    "結案數": row.get("closed_count", 0),
                    "未結案數": row.get("open_count", 0),
                    "結案率(%)": row.get(
                        "close_rate_pct",
                        round(
                            (row.get("closed_count", 0) * 100.0 / row.get("total_count", 1))
                            if row.get("total_count")
                            else 0.0,
                            1,
                        ),
                    ),
                }
                for idx, row in enumerate(resp_stats, start=1)
            ],
            columns=["排名", "責任人", "異常數", "結案數", "未結案數", "結案率(%)"],
        )
        detail_df = pd.DataFrame(
            [
                {
                    "日期": row["event_date"],
                    "類型": "異常" if row["event_type"] == "ANOMALY" else "訪廠",
                    "責任人": str(row.get("responsible_person") or "").strip() or "未指定",
                    "供應商": row["supplier_name"],
                    "問題/摘要": row["content"],
                    "狀態": row["status"],
                }
                for row in rows
            ]
        )
        output = Path(path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="月統計", index=False)
            resp_ranking_df.to_excel(writer, sheet_name="責任人排行", index=False)
            ranking_df.to_excel(writer, sheet_name="供應商排行", index=False)
            detail_df.to_excel(writer, sheet_name="明細", index=False)
        return True, f"已匯出至：{output}"
    except Exception as exc:
        logger.exception("月報匯出失敗")
        return False, f"匯出失敗：{exc}"


def export_events_report(
    file_path: str,
    start_date: str,
    end_date: str,
    temp_chart_paths: dict[str, str] | None = None,
) -> tuple[bool, str]:
    """匯出格式優化後的供應商異常事件統計分析報告，包含視覺總覽與明細/排行表格。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from datetime import datetime

    try:
        events = _query_service.list_events_by_range(start_date, end_date)

        totals, ranking_rows = _query_service.summarize_range_events(events)
        category_pareto_rows = _query_service.get_anomaly_category_pareto_by_range(start_date, end_date)
        process_keyword_pareto_rows = (
            _query_service.get_anomaly_process_keyword_pareto_by_range(start_date, end_date)
        )
        total_anomalies = totals["total_anomalies"]
        closed_anomalies = totals["closed_anomalies"]
        open_anomalies = totals["open_anomalies"]
        close_rate = totals["close_rate"]
        supplier_coverage = totals["supplier_coverage"]
        closure_activity_count = _query_service.get_anomaly_closure_activity_by_range(
            start_date, end_date
        )

        workbook = Workbook()

        try:
            from services.appearance_preferences_service import load_application_preferences
            prefs = load_application_preferences()
        except Exception:
            prefs = None

        excel_theme = getattr(prefs, "excel_theme_style", "classic_navy")
        if excel_theme == "slate_gray":
            primary_color = "334155"
            kpi_bg_color = "F1F5F9"
        elif excel_theme == "forest_green":
            primary_color = "14532D"
            kpi_bg_color = "F0FDF4"
        else:
            primary_color = "1E3A8A"
            kpi_bg_color = "EFF6FF"

        # 樣式定義
        FONT_NAME = "Microsoft JhengHei"
        STYLE_FONT = Font(name=FONT_NAME, size=11)
        STYLE_FONT_BOLD = Font(name=FONT_NAME, size=11, bold=True)
        STYLE_HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        STYLE_TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color=primary_color)
        STYLE_SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")

        STYLE_FILL_HEADER = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
        STYLE_FILL_ZEBRA = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        STYLE_FILL_KPI_BG = PatternFill(start_color=kpi_bg_color, end_color=kpi_bg_color, fill_type="solid")
        STYLE_FILL_TOTAL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        STYLE_BORDER_THIN = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )
        STYLE_BORDER_TOTAL = Border(
            top=Side(style="thin", color="9CA3AF"),
            bottom=Side(style="double", color="111827")
        )

        ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
        ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

        def _auto_fit(ws):
            for col in ws.columns:
                vals = [str(c.value or "") for c in col]
                max_len = max((len(v.encode('utf-8')) for v in vals), default=0)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

        # 1. 視覺總覽報告頁
        report_sheet = workbook.active
        report_sheet.title = "統計報告"
        report_sheet.views.sheetView[0].showGridLines = True

        report_sheet.cell(row=1, column=1, value="供應商品質異常事件統計分析報告").font = STYLE_TITLE_FONT
        report_sheet.row_dimensions[1].height = 30

        subtitle_text = f"統計區間：{start_date} 至 {end_date}   |   報告生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        report_sheet.cell(row=2, column=1, value=subtitle_text).font = STYLE_SUBTITLE_FONT

        # KPI 卡片一 (A4:C5)
        report_sheet.merge_cells("A4:C4")
        report_sheet.cell(row=4, column=1, value="📊 異常事件統計總覽").font = STYLE_FONT_BOLD
        report_sheet.cell(row=4, column=1).alignment = ALIGN_CENTER
        report_sheet.cell(row=4, column=1).fill = STYLE_FILL_KPI_BG

        report_sheet.cell(row=5, column=1, value=f"期間新增異常: {total_anomalies} 件").font = STYLE_FONT
        report_sheet.cell(row=5, column=1).alignment = ALIGN_CENTER
        report_sheet.merge_cells("B5:C5")
        report_sheet.cell(row=5, column=2, value=f"供應商涵蓋: {supplier_coverage} 家").font = STYLE_FONT
        report_sheet.cell(row=5, column=2).alignment = ALIGN_CENTER

        # KPI 卡片二 (E4:G5)
        report_sheet.merge_cells("E4:G4")
        report_sheet.cell(row=4, column=5, value="🎯 處理績效與涵蓋率").font = STYLE_FONT_BOLD
        report_sheet.cell(row=4, column=5).alignment = ALIGN_CENTER
        report_sheet.cell(row=4, column=5).fill = STYLE_FILL_KPI_BG

        report_sheet.cell(row=5, column=5, value=f"期間新增異常－其中目前已結案/未結案: {closed_anomalies} / {open_anomalies}").font = STYLE_FONT
        report_sheet.cell(row=5, column=5).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=6, value=f"期間新增異常目前結案率: {close_rate:.1f}%").font = STYLE_FONT
        report_sheet.cell(row=5, column=6).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=7, value=f"期間結案件數（依結案日期）: {closure_activity_count} 件；供應商 {supplier_coverage} 家").font = STYLE_FONT
        report_sheet.cell(row=5, column=7).alignment = ALIGN_CENTER

        for r in [4, 5]:
            for c in [1, 2, 3, 5, 6, 7]:
                report_sheet.cell(row=r, column=c).border = STYLE_BORDER_THIN

        # 插入統計圖表 (橫向與縱向並排)
        include_charts = bool(getattr(prefs, "export_include_charts", True))
        if temp_chart_paths and include_charts:
            chart_placements = [
                ("trend", "A7"),
                ("category_pareto", "I7"),
                ("responsible", "A23"),
                ("process_keyword_pareto", "A39"),
            ]
            for key, cell in chart_placements:
                path = temp_chart_paths.get(key)
                if path and Path(path).exists():
                    img = Image(path)
                    img.width = 460
                    img.height = 310
                    report_sheet.add_image(img, cell)

        # 2. 異常類別柏拉圖資料頁
        category_sheet = workbook.create_sheet("異常類別柏拉圖")
        category_sheet.views.sheetView[0].showGridLines = True

        category_headers = ["排名", "異常類別", "件數", "佔比(%)", "累積佔比(%)"]
        category_sheet.append(category_headers)
        for col_idx in range(1, len(category_headers) + 1):
            cell = category_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        category_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(category_pareto_rows, start=2):
            data = [
                row.get("rank", 0),
                row.get("category", ""),
                row.get("count", 0),
                row.get("percent", 0.0),
                row.get("cumulative_percent", 0.0),
            ]
            category_sheet.append(data)

            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(category_headers) + 1):
                cell = category_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA
                if c_idx == 2:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
                if c_idx in (4, 5):
                    cell.number_format = "0.0"
            category_sheet.row_dimensions[r_idx].height = 20
        _auto_fit(category_sheet)

        keyword_sheet = workbook.create_sheet("SMT製程關鍵詞柏拉圖")
        keyword_sheet.views.sheetView[0].showGridLines = True
        keyword_headers = ["排名", "關鍵詞", "件數", "佔比(%)", "累積佔比(%)"]
        keyword_sheet.append(keyword_headers)
        for col_idx in range(1, len(keyword_headers) + 1):
            cell = keyword_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        keyword_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(process_keyword_pareto_rows, start=2):
            data = [
                row.get("rank", 0),
                row.get("keyword", ""),
                row.get("count", 0),
                row.get("percent", 0.0),
                row.get("cumulative_percent", 0.0),
            ]
            keyword_sheet.append(data)
            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(keyword_headers) + 1):
                cell = keyword_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA
                if c_idx == 2:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
                if c_idx in (4, 5):
                    cell.number_format = "0.0"
            keyword_sheet.row_dimensions[r_idx].height = 20
        _auto_fit(keyword_sheet)

        # 3. 事件明細依權威 event_type 分成訪廠與異常兩個活頁。
        def _build_event_detail_sheet(name, headers, rows, row_builder, centered_columns):
            sheet = workbook.create_sheet(name)
            sheet.views.sheetView[0].showGridLines = True
            sheet.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.font = STYLE_HEADER_FONT
                cell.fill = STYLE_FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = STYLE_BORDER_THIN
            sheet.row_dimensions[1].height = 24

            for r_idx, row in enumerate(rows, start=2):
                sheet.append(row_builder(row))
                is_even = r_idx % 2 == 0
                for c_idx in range(1, len(headers) + 1):
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    cell.font = STYLE_FONT
                    cell.border = STYLE_BORDER_THIN
                    if is_even:
                        cell.fill = STYLE_FILL_ZEBRA
                    cell.alignment = ALIGN_CENTER if c_idx in centered_columns else ALIGN_LEFT
                sheet.row_dimensions[r_idx].height = 20
            _auto_fit(sheet)

        anomaly_rows = [row for row in events if row.get("event_type") == "ANOMALY"]

        def _anomaly_detail_row(row):
            from services.process_keyword_codec import format_process_keywords_display

            if row.get("quality_report_required") is None:
                quality_report_required = "未設定"
            else:
                quality_report_required = "是" if bool(row.get("quality_report_required")) else "否"
            current = row.get("current_action") or {}
            current_desc = str(current.get("description") or "").strip()
            current_owner = str(current.get("owner") or "").strip()
            current_due = str(current.get("due_date") or "").strip()
            if current_desc:
                if current_due:
                    current_text = f"{current_desc}（{current_owner or '—'} / {current_due}）"
                else:
                    current_text = f"{current_desc}（{current_owner or '—'}）"
            else:
                current_text = ""
            return [
                str(row.get("ref_no") or ""),
                row.get("event_date") or "",
                row.get("supplier_name") or "",
                row.get("product_code") or "",
                row.get("product_name") or "",
                row.get("product_stage") or "",
                row.get("anomaly_source") or "",
                row.get("material_receipt_no") or "",
                row.get("internal_work_order_no") or "",
                row.get("outsource_work_order") or row.get("work_order_no") or "",
                row.get("outsource_receipt_no") or "",
                row.get("category") or "",
                format_process_keywords_display(row.get("process_keywords")),
                str(row.get("responsible_person") or "").strip() or "未指定",
                row.get("content") or "",
                row.get("pending_items") or "",
                quality_report_required,
                row.get("improvement_desc") or "",
                "是" if row.get("overdue") else "否",
                current_text,
                int(row.get("open_action_count") or 0),
                str(row.get("root_cause_status") or "尚未開始"),
                str(row.get("corrective_action_status") or "—"),
                str(row.get("verification_result") or "—"),
                int(row.get("attachment_count") or 0),
                int(row.get("hypothesis_count") or 0),
                "是" if row.get("hypothesis_adopted") else "否",
                int(row.get("repeat_link_count") or 0),
                row.get("status") or "",
                row.get("closed_at") or "",
            ]

        _build_event_detail_sheet(
            "異常",
            [
                "異常單號",
                "日期",
                "供應商",
                "料號",
                "品名",
                "階段",
                "異常來源",
                "原物料進貨單號",
                "廠內製令單號",
                "委外製令單號",
                "委外進貨單號",
                "異常類別",
                "SMT 製程關鍵詞",
                "責任人",
                "問題/摘要",
                "確認事項 / 待追蹤",
                "品質異常單要求",
                "改善說明",
                _OVERVIEW_LABELS["overdue"],
                _OVERVIEW_LABELS["current_action_text"],
                _OVERVIEW_LABELS["open_action_count"],
                _OVERVIEW_LABELS["root_cause_status"],
                _OVERVIEW_LABELS["corrective_action_status"],
                _OVERVIEW_LABELS["verification_result"],
                _OVERVIEW_LABELS["attachment_count"],
                _OVERVIEW_LABELS["hypothesis_count"],
                _OVERVIEW_LABELS["hypothesis_adopted"],
                _OVERVIEW_LABELS["repeat_link_count"],
                "狀態",
                "結案日期",
            ],
            anomaly_rows,
            _anomaly_detail_row,
            {1, 2, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30},
        )

        export_warnings: list[str] = []
        hypothesis_temp_pngs: list[Path] = []
        if include_charts:
            hypothesis_temp_pngs = _append_hypothesis_export_sheet(
                workbook,
                anomaly_rows,
                output_parent=Path(file_path).parent,
                warnings=export_warnings,
            )

        # 4. 責任人排行榜頁
        resp_stats_rows = _query_service.get_responsible_person_stats_by_range(start_date, end_date)
        resp_sheet = workbook.create_sheet("責任人排行榜")
        resp_sheet.views.sheetView[0].showGridLines = True

        resp_headers = ["排名", "責任人", "期間新增異常", "其中目前已結案", "其中目前未結案", "目前結案率(%)"]
        resp_sheet.append(resp_headers)
        for col_idx in range(1, len(resp_headers) + 1):
            cell = resp_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        resp_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(resp_stats_rows, start=2):
            data = [
                r_idx - 1,
                row.get("responsible_person", "未指定"),
                row.get("anomaly_count", 0),
                row.get("closed_count", 0),
                row.get("open_count", 0),
                f"{row.get('close_rate_pct', 0.0):.1f}%"
            ]
            resp_sheet.append(data)

            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(resp_headers) + 1):
                cell = resp_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA

                if c_idx == 1:
                    cell.alignment = ALIGN_CENTER
                elif c_idx == 2:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
            resp_sheet.row_dimensions[r_idx].height = 20

        total_resp_row_idx = len(resp_stats_rows) + 2
        if len(resp_stats_rows) > 0:
            resp_sheet.cell(row=total_resp_row_idx, column=1, value="合計").font = STYLE_FONT_BOLD
            resp_sheet.cell(row=total_resp_row_idx, column=1).alignment = ALIGN_CENTER

            for c_idx in (3, 4, 5):
                col_letter = get_column_letter(c_idx)
                sum_formula = f"=SUM({col_letter}2:{col_letter}{total_resp_row_idx - 1})"
                cell = resp_sheet.cell(row=total_resp_row_idx, column=c_idx, value=sum_formula)
                cell.font = STYLE_FONT_BOLD
                cell.alignment = ALIGN_RIGHT
                cell.number_format = "#,##0"

            total_closed_cell = f"D{total_resp_row_idx}"
            total_anomaly_cell = f"C{total_resp_row_idx}"
            rate_formula = f"=IF({total_anomaly_cell}>0, {total_closed_cell}/{total_anomaly_cell}, 0)"
            rate_cell = resp_sheet.cell(row=total_resp_row_idx, column=6, value=rate_formula)
            rate_cell.font = STYLE_FONT_BOLD
            rate_cell.alignment = ALIGN_RIGHT
            rate_cell.number_format = "0.0%"

            for c_idx in range(1, len(resp_headers) + 1):
                cell = resp_sheet.cell(row=total_resp_row_idx, column=c_idx)
                cell.border = STYLE_BORDER_TOTAL
                cell.fill = STYLE_FILL_TOTAL

        _auto_fit(resp_sheet)

        # 5. 供應商排行榜頁
        rank_sheet = workbook.create_sheet("供應商排行榜")
        rank_sheet.views.sheetView[0].showGridLines = True

        rank_headers = ["排名", "供應商", "期間新增異常", "其中目前已結案", "其中目前未結案", "目前結案率(%)"]
        rank_sheet.append(rank_headers)
        for col_idx in range(1, len(rank_headers) + 1):
            cell = rank_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        rank_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(ranking_rows, start=2):
            data = [
                r_idx - 1,
                row.get("supplier_name", ""),
                row.get("anomaly_count", 0),
                row.get("closed_anomaly_count", 0),
                row.get("open_anomaly_count", 0),
                f"{row.get('close_rate_pct', 0.0):.1f}%"
            ]
            rank_sheet.append(data)

            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(rank_headers) + 1):
                cell = rank_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA

                if c_idx == 1:
                    cell.alignment = ALIGN_CENTER
                elif c_idx == 2:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
            rank_sheet.row_dimensions[r_idx].height = 20

        # 合計列
        total_row_idx = len(ranking_rows) + 2
        if len(ranking_rows) > 0:
            rank_sheet.cell(row=total_row_idx, column=1, value="合計").font = STYLE_FONT_BOLD
            rank_sheet.cell(row=total_row_idx, column=1).alignment = ALIGN_CENTER

            for c_idx in (3, 4, 5):
                col_letter = get_column_letter(c_idx)
                sum_formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                cell = rank_sheet.cell(row=total_row_idx, column=c_idx, value=sum_formula)
                cell.font = STYLE_FONT_BOLD
                cell.alignment = ALIGN_RIGHT
                cell.number_format = "#,##0"

            # 總體結案率公式 = 總已結案數 / 總異常件數
            total_closed_cell = f"D{total_row_idx}"
            total_anomaly_cell = f"C{total_row_idx}"
            rate_formula = f"=IF({total_anomaly_cell}>0, {total_closed_cell}/{total_anomaly_cell}, 0)"
            rate_cell = rank_sheet.cell(row=total_row_idx, column=6, value=rate_formula)
            rate_cell.font = STYLE_FONT_BOLD
            rate_cell.alignment = ALIGN_RIGHT
            rate_cell.number_format = "0.0%"

            for c_idx in range(1, len(rank_headers) + 1):
                cell = rank_sheet.cell(row=total_row_idx, column=c_idx)
                cell.border = STYLE_BORDER_TOTAL
                cell.fill = STYLE_FILL_TOTAL

            if getattr(prefs, "export_include_disclaimer", True):
                disclaimer_text = "※ 免責聲明：本報表所載之品質資料屬 Mitcorp 內部機密，僅供 SQE 供應商管理與工程分析使用，未經授權禁止外傳。"
                disclaimer_row = total_row_idx + 2
                rank_sheet.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=len(rank_headers))
                d_cell = rank_sheet.cell(row=disclaimer_row, column=1, value=disclaimer_text)
                d_cell.font = STYLE_SUBTITLE_FONT
                d_cell.alignment = ALIGN_LEFT

        _auto_fit(rank_sheet)

        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        for png_path in hypothesis_temp_pngs:
            try:
                png_path.unlink(missing_ok=True)
            except OSError:
                pass
        message = f"已匯出至：{output_path}"
        if export_warnings:
            message += "\n完成但有警告：\n- " + "\n- ".join(export_warnings)
        return True, message
    except Exception as exc:
        logger.exception("自訂日期區間 Excel 報告匯出出錯")
        return False, f"匯出報告失敗：{exc}"
