"""Service layer for v2 minimalist events workflow."""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING

logger = logging.getLogger(__name__)

from database.connection import get_connection
from database import repository
from services import event_pdf_exporter

if TYPE_CHECKING:
    from PySide6.QtGui import QImage

EVENT_SCOPE_VISIT_ONLY = repository.EVENT_SCOPE_VISIT_ONLY
EVENT_SCOPE_VISIT_WITH_ANOMALY = repository.EVENT_SCOPE_VISIT_WITH_ANOMALY
EVENT_SCOPE_ANOMALY_ONLY = repository.EVENT_SCOPE_ANOMALY_ONLY
EVENT_SCOPE_CLOSED_ONLY = repository.EVENT_SCOPE_CLOSED_ONLY


def _month_now() -> str:
    return date.today().strftime("%Y%m")


def _require_supplier_record(
    conn,
    supplier_id: str,
    *,
    require_active: bool,
) -> dict:
    supplier_key = (supplier_id or "").strip()
    if not supplier_key:
        raise ValueError("Supplier is required")
    supplier = repository.get_supplier(conn, supplier_key)
    if supplier is None:
        raise ValueError("Supplier not found")
    if require_active and not bool(supplier.get("is_active")):
        raise ValueError("Supplier is inactive")
    return supplier


def _product_matches_supplier_scope(product: dict, supplier_id: str) -> bool:
    primary = str(product.get("supplier_id") or "").strip()
    secondary = str(product.get("secondary_supplier_id") or "").strip()
    if primary == supplier_id or secondary == supplier_id:
        return True
    if not primary and not secondary:
        return True
    return False


def _resolve_product_name(
    conn,
    *,
    supplier_id: str,
    product_id: str | None,
    require_active: bool = False,
) -> str:
    product_key = (product_id or "").strip()
    if not product_key:
        return ""
    product = repository.get_product(conn, product_key)
    if product is None:
        raise ValueError("Product not found")
    if require_active and not bool(product.get("is_active")):
        raise ValueError("Product is inactive")
    if not _product_matches_supplier_scope(product, supplier_id):
        raise ValueError("Product does not belong to selected supplier")
    return str(product.get("product_name") or "").strip()


def _require_product_id(payload: dict) -> str:
    product_id = (payload.get("product_id") or "").strip()
    if not product_id:
        raise ValueError("Product is required")
    return product_id


def _has_visit_record_content(payload: dict) -> bool:
    if (payload.get("product_id") or "").strip():
        return True
    for section in payload.get("product_sections") or []:
        item = dict(section or {})
        if any(
            (
                (item.get("product_id") or "").strip(),
                (item.get("product_name") or "").strip(),
                (item.get("time_slot") or "").strip(),
                (item.get("work_order_no") or "").strip(),
                (item.get("summary") or "").strip(),
                str(item.get("production_qty") or "").strip() not in {"", "0"},
                _visit_note_has_content(item.get("defect_notes")),
            )
        ):
            return True
    if _visit_note_has_content(payload.get("defect_notes")):
        return True
    return False


def _visit_note_has_content(notes: object) -> bool:
    for note in notes or []:
        item = dict(note or {})
        if any(
            str(item.get(key) or "").strip()
            for key in ("defect_desc", "defect", "description", "improvement_desc", "note", "remark")
        ):
            return True
    return False


def _validate_visit_product_scope(
    conn,
    *,
    supplier_id: str,
    payload: dict,
    require_active: bool,
) -> None:
    product_id = (payload.get("product_id") or "").strip()
    if product_id:
        _resolve_product_name(
            conn,
            supplier_id=supplier_id,
            product_id=product_id,
            require_active=require_active,
        )
    for section in payload.get("product_sections") or []:
        section_id = (dict(section or {}).get("product_id") or "").strip()
        if not section_id:
            continue
        _resolve_product_name(
            conn,
            supplier_id=supplier_id,
            product_id=section_id,
            require_active=require_active,
        )


def list_suppliers(*, include_inactive: bool = True) -> list[dict]:
    with get_connection() as conn:
        return repository.list_suppliers(conn, include_inactive=include_inactive)


def create_supplier(payload: dict) -> str:
    with get_connection() as conn:
        return repository.create_supplier_record(
            conn,
            supplier_name=payload.get("supplier_name", ""),
            contact_name=payload.get("contact_name", ""),
            department=payload.get("department", ""),
            phone=payload.get("phone", ""),
            contact_email=payload.get("contact_email", ""),
        )


def update_supplier(supplier_id: str, payload: dict) -> None:
    with get_connection() as conn:
        repository.update_supplier_record(
            conn,
            supplier_id=supplier_id,
            supplier_name=payload.get("supplier_name", ""),
            contact_name=payload.get("contact_name", ""),
            department=payload.get("department", ""),
            phone=payload.get("phone", ""),
            contact_email=payload.get("contact_email", ""),
        )


def set_supplier_active(supplier_id: str, is_active: bool) -> None:
    with get_connection() as conn:
        repository.set_supplier_active(conn, supplier_id, is_active)


def delete_supplier(supplier_id: str) -> None:
    with get_connection() as conn:
        repository.delete_supplier_record(conn, supplier_id)


def list_supplier_contacts(supplier_id: str) -> list[dict]:
    with get_connection() as conn:
        return repository.list_supplier_contacts(conn, supplier_id)


def add_supplier_contact(supplier_id: str, payload: dict) -> str:
    with get_connection() as conn:
        return repository.add_supplier_contact(
            conn,
            supplier_id=supplier_id,
            contact_name=payload.get("contact_name", ""),
            department=payload.get("department", ""),
            phone=payload.get("phone", ""),
            email=payload.get("email", ""),
            is_primary=bool(payload.get("is_primary", False)),
        )


def delete_supplier_contact(contact_id: str) -> None:
    with get_connection() as conn:
        repository.delete_supplier_contact(conn, contact_id)


def set_primary_contact(supplier_id: str, contact_id: str) -> None:
    with get_connection() as conn:
        repository.set_primary_contact(conn, supplier_id, contact_id)


def delete_suppliers(
    supplier_ids: list[str],
) -> repository.SupplierDeleteResult:
    with get_connection() as conn:
        return repository.delete_supplier_records(conn, supplier_ids)


def list_products(*, include_inactive: bool = True) -> list[dict]:
    with get_connection() as conn:
        return repository.list_products(conn, include_inactive=include_inactive)


def create_product(payload: dict) -> str:
    with get_connection() as conn:
        return repository.create_product_record(
            conn,
            product_code=payload.get("product_code", ""),
            product_name=payload.get("product_name", ""),
            product_stage=payload.get("product_stage", "量產"),
            supplier_id=(payload.get("supplier_id") or "").strip(),
            secondary_supplier_id=(
                payload.get("secondary_supplier_id") or ""
            ).strip()
            or None,
        )


def update_product(product_id: str, payload: dict) -> None:
    with get_connection() as conn:
        repository.update_product_record(
            conn,
            product_id=product_id,
            product_code=payload.get("product_code", ""),
            product_name=payload.get("product_name", ""),
            product_stage=payload.get("product_stage", "量產"),
            supplier_id=(payload.get("supplier_id") or "").strip(),
            secondary_supplier_id=(
                payload.get("secondary_supplier_id") or ""
            ).strip()
            or None,
            stage_change_reason=(payload.get("stage_change_reason") or "").strip(),
        )


def set_product_active(product_id: str, is_active: bool) -> None:
    with get_connection() as conn:
        repository.set_product_active(conn, product_id, is_active)


def delete_product(product_id: str) -> None:
    with get_connection() as conn:
        repository.delete_product_record(conn, product_id)


def list_active_suppliers() -> list[dict]:
    with get_connection() as conn:
        return repository.list_active_suppliers(conn)


def list_active_products_for_supplier(supplier_id: str | None) -> list[dict]:
    with get_connection() as conn:
        return repository.list_active_products_for_supplier(conn, supplier_id)


def has_active_suppliers() -> bool:
    return bool(list_active_suppliers())


def create_anomaly(payload: dict) -> str:
    problem_desc = (payload.get("problem_desc") or "").strip()
    if not problem_desc:
        raise ValueError("Problem description is required")

    supplier_id = (payload.get("supplier_id") or "").strip()
    product_id = _require_product_id(payload)
    anomaly_date = payload.get("anomaly_date") or date.today().isoformat()

    with get_connection() as conn:
        _require_supplier_record(conn, supplier_id, require_active=True)
        product_name = _resolve_product_name(
            conn,
            supplier_id=supplier_id,
            product_id=product_id,
            require_active=True,
        )
        return repository.create_anomaly(
            conn,
            anomaly_date=anomaly_date,
            supplier_id=supplier_id,
            problem_desc=problem_desc,
            category=payload.get("category", ""),
            product_lot_no=payload.get("product_lot_no", ""),
            product_id=product_id,
            product_name=product_name,
            outsource_work_order=payload.get("outsource_work_order", ""),
            batch_qty=payload.get("batch_qty", 0),
            pending_items=payload.get("pending_items", ""),
            responsible_person=payload.get("responsible_person", ""),
            due_date=payload.get("due_date", ""),
            rc_supplier_inventory=payload.get("rc_supplier_inventory", "unconfirmed"),
            rc_supplier_wip=payload.get("rc_supplier_wip", "unconfirmed"),
            rc_in_transit=payload.get("rc_in_transit", "unconfirmed"),
            rc_internal_inventory=payload.get("rc_internal_inventory", "unconfirmed"),
            is_tech_transfer=bool(payload.get("is_tech_transfer", False)),
        )


def create_anomaly_with_visit_link(payload: dict) -> dict:
    problem_desc = (payload.get("problem_desc") or "").strip()
    if not problem_desc:
        raise ValueError("Problem description is required")

    supplier_id = (payload.get("supplier_id") or "").strip()
    product_id = _require_product_id(payload)
    anomaly_date = payload.get("anomaly_date") or date.today().isoformat()
    visit_id = (payload.get("visit_id") or "").strip() or None
    sync_visit = bool(payload.get("sync_visit", True))
    visit_summary = payload.get("visit_summary", "")

    with get_connection() as conn:
        _require_supplier_record(conn, supplier_id, require_active=True)
        product_name = _resolve_product_name(
            conn,
            supplier_id=supplier_id,
            product_id=product_id,
            require_active=True,
        )
        return repository.create_anomaly_with_visit_link(
            conn,
            anomaly_date=anomaly_date,
            supplier_id=supplier_id,
            problem_desc=problem_desc,
            category=payload.get("category", ""),
            product_lot_no=payload.get("product_lot_no", ""),
            product_id=product_id,
            product_name=product_name,
            outsource_work_order=payload.get("outsource_work_order", ""),
            batch_qty=payload.get("batch_qty", 0),
            visit_id=visit_id,
            sync_visit=sync_visit,
            visit_summary=visit_summary,
            pending_items=payload.get("pending_items", ""),
            responsible_person=payload.get("responsible_person", ""),
            due_date=payload.get("due_date", ""),
            rc_supplier_inventory=payload.get("rc_supplier_inventory", "unconfirmed"),
            rc_supplier_wip=payload.get("rc_supplier_wip", "unconfirmed"),
            rc_in_transit=payload.get("rc_in_transit", "unconfirmed"),
            rc_internal_inventory=payload.get("rc_internal_inventory", "unconfirmed"),
            is_tech_transfer=bool(payload.get("is_tech_transfer", False)),
            anomaly_no=payload.get("anomaly_no"),
        )


def get_anomaly_detail(anomaly_id: str) -> dict:
    if not (anomaly_id or "").strip():
        raise ValueError("Anomaly id is required")
    with get_connection() as conn:
        row = repository.get_anomaly_detail(conn, anomaly_id)
    if row is None:
        raise ValueError("Anomaly not found")
    return row


def update_anomaly(anomaly_id: str, payload: dict) -> None:
    anomaly_key = (anomaly_id or "").strip()
    if not anomaly_key:
        raise ValueError("Anomaly id is required")

    supplier_id = (payload.get("supplier_id") or "").strip()
    product_id = _require_product_id(payload)
    anomaly_date = payload.get("anomaly_date") or date.today().isoformat()
    problem_desc = (payload.get("problem_desc") or "").strip()
    if not problem_desc:
        raise ValueError("Problem description is required")

    with get_connection() as conn:
        _require_supplier_record(conn, supplier_id, require_active=False)
        product_name = _resolve_product_name(
            conn,
            supplier_id=supplier_id,
            product_id=product_id,
        )
        repository.update_anomaly(
            conn,
            anomaly_id=anomaly_key,
            anomaly_date=anomaly_date,
            supplier_id=supplier_id,
            problem_desc=problem_desc,
            category=payload.get("category", ""),
            product_lot_no=payload.get("product_lot_no", ""),
            product_id=product_id,
            product_name=product_name,
            outsource_work_order=payload.get("outsource_work_order", ""),
            batch_qty=payload.get("batch_qty", 0),
            pending_items=payload.get("pending_items", ""),
            responsible_person=payload.get("responsible_person", ""),
            due_date=payload.get("due_date", ""),
            rc_supplier_inventory=payload.get("rc_supplier_inventory", "unconfirmed"),
            rc_supplier_wip=payload.get("rc_supplier_wip", "unconfirmed"),
            rc_in_transit=payload.get("rc_in_transit", "unconfirmed"),
            rc_internal_inventory=payload.get("rc_internal_inventory", "unconfirmed"),
            is_tech_transfer=bool(payload.get("is_tech_transfer", False)),
            anomaly_no=payload.get("anomaly_no"),
        )
        conn.commit()


def update_anomaly_link(anomaly_id: str, visit_id: str | None) -> None:
    """Manually update the visit association for an existing anomaly."""
    if not (anomaly_id or "").strip():
        raise ValueError("Anomaly id is required")
    with get_connection() as conn:
        repository.update_anomaly_link(conn, anomaly_id, visit_id)
        conn.commit()


def list_visits_for_supplier(supplier_id: str) -> list[dict]:
    """Return all visit records for a specific supplier, ordered by date."""
    if not (supplier_id or "").strip():
        return []
    with get_connection() as conn:
        return repository.list_visits_by_supplier(conn, supplier_id)


def list_pending_visit_defect_notes(*, limit: int | None = None) -> list[dict]:
    with get_connection() as conn:
        return repository.list_pending_visit_defect_notes(conn, limit=limit)


def confirm_visit_defect_note_as_anomaly(note_id: str, payload: dict | None = None) -> dict:
    params = payload or {}
    with get_connection() as conn:
        return repository.confirm_visit_defect_note_as_anomaly(
            conn,
            note_id=note_id,
            product_id=params.get("product_id"),
            responsible_person=params.get("responsible_person", ""),
            due_date=params.get("due_date", ""),
        )


def delete_anomaly(anomaly_id: str) -> None:
    anomaly_key = (anomaly_id or "").strip()
    if not anomaly_key:
        raise ValueError("Anomaly id is required")
    with get_connection() as conn:
        repository.delete_anomaly(conn, anomaly_key)


def preview_anomaly_no(anomaly_date: str | None = None) -> str:
    target_date = anomaly_date or date.today().isoformat()
    with get_connection() as conn:
        return repository.preview_anomaly_no(conn, target_date)


def get_latest_tech_transfer_for_supplier(supplier_id: str) -> dict | None:
    """查詢指定供應商最新一筆含技轉資料的訪廠紀錄，做為新增異常的「參考資料」。
    若查無技轉紀錄則回傳 None。"""
    normalized = (supplier_id or "").strip()
    if not normalized:
        return None
    with get_connection() as conn:
        return repository.get_latest_tech_transfer_for_supplier(conn, normalized)


def get_latest_visit_for_supplier_on_date(
    supplier_id: str, visit_date: str
) -> dict | None:
    normalized_supplier = (supplier_id or "").strip()
    normalized_date = (visit_date or "").strip()
    if not normalized_supplier or not normalized_date:
        return None
    with get_connection() as conn:
        return repository.get_latest_visit_for_supplier_on_date(
            conn,
            supplier_id=normalized_supplier,
            visit_date=normalized_date,
        )


def close_anomaly(
    anomaly_id: str,
    improvement_desc: str,
    *,
    closed_by: str = "",
    root_cause_category: str = "",
) -> None:
    if not (anomaly_id or "").strip():
        raise ValueError("Anomaly id is required")
    text = (improvement_desc or "").strip()
    if not text:
        raise ValueError("Improvement description is required")
    with get_connection() as conn:
        repository.close_anomaly(
            conn,
            anomaly_id=anomaly_id,
            improvement_desc=improvement_desc,
            closed_by=closed_by,
            root_cause_category=root_cause_category,
        )


def reopen_anomaly(anomaly_id: str) -> None:
    if not (anomaly_id or "").strip():
        raise ValueError("Anomaly id is required")
    with get_connection() as conn:
        repository.reopen_anomaly(conn, anomaly_id)


def create_visit(payload: dict) -> str:
    supplier_id = (payload.get("supplier_id") or "").strip()
    if not _has_visit_record_content(payload):
        _require_product_id(payload)
    product_id = (payload.get("product_id") or "").strip()
    visit_date = payload.get("visit_date") or date.today().isoformat()
    with get_connection() as conn:
        _require_supplier_record(conn, supplier_id, require_active=True)
        _validate_visit_product_scope(
            conn,
            supplier_id=supplier_id,
            payload=payload,
            require_active=True,
        )
        product_name = (
            _resolve_product_name(
                conn,
                supplier_id=supplier_id,
                product_id=product_id,
                require_active=True,
            )
            if product_id
            else ""
        )
        return repository.create_visit(
            conn,
            visit_date=visit_date,
            supplier_id=supplier_id,
            product_id=product_id,
            product_name=product_name,
            visitor_name=payload.get("visitor_name", ""),
            summary=payload.get("summary", ""),
            work_order_no=payload.get("work_order_no", ""),
            production_qty=payload.get("production_qty", 0),
            product_sections=payload.get("product_sections"),
            defect_notes=payload.get("defect_notes"),
            tech_transfer=bool(payload.get("tech_transfer", False)),
            tech_transfer_doc=bool(payload.get("tech_transfer_doc", False)),
            carrier_requirement=bool(payload.get("carrier_requirement", False)),
            dispensing_process=bool(payload.get("dispensing_process", False)),
            functional_test=bool(payload.get("functional_test", False)),
            packaging_requirement=bool(payload.get("packaging_requirement", False)),
            tech_transfer_states=payload.get("tech_transfer_states"),
        )


def update_visit(visit_id: str, payload: dict) -> None:
    visit_key = (visit_id or "").strip()
    if not visit_key:
        raise ValueError("Visit id is required")

    supplier_id = (payload.get("supplier_id") or "").strip()
    if not _has_visit_record_content(payload):
        _require_product_id(payload)
    product_id = (payload.get("product_id") or "").strip()
    visit_date = payload.get("visit_date") or date.today().isoformat()

    with get_connection() as conn:
        _require_supplier_record(conn, supplier_id, require_active=False)
        _validate_visit_product_scope(
            conn,
            supplier_id=supplier_id,
            payload=payload,
            require_active=False,
        )
        product_name = (
            _resolve_product_name(
                conn,
                supplier_id=supplier_id,
                product_id=product_id,
            )
            if product_id
            else ""
        )
        repository.update_visit(
            conn,
            visit_id=visit_key,
            visit_date=visit_date,
            supplier_id=supplier_id,
            product_id=product_id,
            product_name=product_name,
            visitor_name=payload.get("visitor_name", ""),
            summary=payload.get("summary", ""),
            work_order_no=payload.get("work_order_no", ""),
            production_qty=payload.get("production_qty", 0),
            product_sections=payload.get("product_sections"),
            defect_notes=payload.get("defect_notes"),
            tech_transfer=bool(payload.get("tech_transfer", False)),
            tech_transfer_doc=bool(payload.get("tech_transfer_doc", False)),
            carrier_requirement=bool(payload.get("carrier_requirement", False)),
            dispensing_process=bool(payload.get("dispensing_process", False)),
            functional_test=bool(payload.get("functional_test", False)),
            packaging_requirement=bool(payload.get("packaging_requirement", False)),
            tech_transfer_states=payload.get("tech_transfer_states"),
        )


def get_visit_detail(visit_id: str) -> dict:
    if not (visit_id or "").strip():
        raise ValueError("Visit id is required")
    with get_connection() as conn:
        row = repository.get_visit_detail(conn, visit_id)
    if row is None:
        raise ValueError("Visit not found")
    return row


def delete_visit(visit_id: str) -> None:
    visit_key = (visit_id or "").strip()
    if not visit_key:
        raise ValueError("Visit id is required")
    with get_connection() as conn:
        repository.delete_visit(conn, visit_key)


def list_events(filters: dict | None = None) -> list[dict]:
    params = filters or {}
    with get_connection() as conn:
        return repository.list_events(
            conn,
            event_type=params.get("event_type", "ALL"),
            status=params.get("status", "ALL"),
            supplier_keyword=params.get("supplier", ""),
            yyyymm=params.get("yyyymm"),
            limit=params.get("limit"),
            event_scope=params.get("event_scope"),
            overdue_only=bool(params.get("overdue_only", False)),
        )


def _event_pdf_payload(row: dict) -> tuple[dict, dict | None]:
    event_id = str(row.get("event_id") or "").strip()
    if not event_id:
        raise ValueError("Event id is required")
    event_type = str(row.get("event_type") or "").strip().upper()
    if event_type == "VISIT":
        return get_visit_detail(event_id), None
    if event_type == "ANOMALY":
        detail = get_anomaly_detail(event_id)
        linked_visit_id = str(
            row.get("linked_visit_id") or detail.get("visit_id") or ""
        ).strip()
        linked_visit = get_visit_detail(linked_visit_id) if linked_visit_id else None
        return detail, linked_visit
    raise ValueError("Event type is required")


def default_event_pdf_filename(row: dict) -> str:
    detail, _linked_visit = _event_pdf_payload(row)
    return event_pdf_exporter.default_event_pdf_filename(row, detail)


def _run_event_pdf_export(
    path: str,
    row: dict,
    delegate,
    *,
    failure_log: str,
    failure_prefix: str,
) -> tuple[bool, str]:
    """Shared payload-resolve → delegate → exception-to-(False, msg) wrapper
    for the full/brief PDF export entry points (audit finding D17)."""
    try:
        detail, linked_visit = _event_pdf_payload(row)
        return delegate(path, row, detail, linked_visit=linked_visit)
    except Exception as exc:
        logger.exception(failure_log)
        return False, f"{failure_prefix}{exc}"


def export_event_pdf(path: str, row: dict) -> tuple[bool, str]:
    return _run_event_pdf_export(
        path,
        row,
        event_pdf_exporter.export_event_pdf,
        failure_log="PDF 匯出失敗",
        failure_prefix="匯出失敗：",
    )


def export_brief_event_pdf(path: str, row: dict) -> tuple[bool, str]:
    return _run_event_pdf_export(
        path,
        row,
        event_pdf_exporter.export_brief_event_pdf,
        failure_log="精簡版 PDF 匯出失敗",
        failure_prefix="匯出精簡版失敗：",
    )


def render_brief_event_image(row: dict) -> "QImage | None":
    """將精簡報告渲染為 QImage 供 LINE 剪貼簿圖片傳送。"""
    try:
        detail, linked_visit = _event_pdf_payload(row)
        return event_pdf_exporter.render_brief_event_to_image(
            row,
            detail,
            linked_visit=linked_visit,
        )
    except Exception:
        logger.exception("渲染精簡報告圖片失敗")
        return None


def get_dashboard_summary() -> dict:
    with get_connection() as conn:
        return repository.get_dashboard_summary(conn)


def get_monthly_stats(yyyymm: str | None = None) -> dict:
    month = yyyymm or _month_now()
    with get_connection() as conn:
        return repository.get_monthly_stats(conn, month)


def get_responsible_person_stats(yyyymm: str | None = None) -> list[dict]:
    month = yyyymm or _month_now()
    with get_connection() as conn:
        return repository.get_responsible_person_stats(conn, month)


def _stats_period_label(yyyymm: str) -> str:
    period_key = str(yyyymm or "").strip().upper()
    today = date.today()
    if period_key == "ALL":
        return "全期項目"
    if period_key == "YEAR":
        return f"{today.year} 年度"
    if period_key == "HALF_YEAR":
        half_label = "上半年" if today.month <= 6 else "下半年"
        return f"{today.year} {half_label}"

    month = str(yyyymm or _month_now()).strip()
    if len(month) == 7 and "-" in month:
        month = month.replace("-", "")
    if len(month) == 6 and month.isdigit():
        return f"{month[:4]}-{month[4:]}"
    return month


def list_product_stage_change_logs(
    *, product_id: str | None = None, limit: int = 200
) -> list[dict]:
    with get_connection() as conn:
        return repository.list_product_stage_change_logs(
            conn,
            product_id=(product_id or "").strip() or None,
            limit=limit,
        )


def export_monthly_excel(path: str, yyyymm: str) -> tuple[bool, str]:
    import pandas as pd
    try:
        month = yyyymm or _month_now()
        stats = get_monthly_stats(month)
        rows = list_events({"yyyymm": month})
        summary_df = pd.DataFrame(
            [
                {
                    "月份": _stats_period_label(month),
                    "本月異常數": stats["anomaly_count"],
                    "訪廠數": stats["visit_count"],
                    "結案數": stats["closed_anomaly_count"],
                    "未結案數": stats["open_anomaly_count"],
                    "結案率(%)": stats["close_rate_pct"],
                    "異常/訪廠比": stats["anomaly_visit_ratio"],
                    "供應商覆蓋數": stats["supplier_coverage_count"],
                }
            ]
        )
        ranking_df = pd.DataFrame(
            [
                {
                    "排名": idx,
                    "供應商": row["supplier_name"],
                    "異常數": row["anomaly_count"],
                    "訪廠數": row["visit_count"],
                    "結案數": row["closed_anomaly_count"],
                    "未結案數": row["open_anomaly_count"],
                    "結案率(%)": row["close_rate_pct"],
                }
                for idx, row in enumerate(stats["top_suppliers_by_anomaly"], start=1)
            ],
            columns=["排名", "供應商", "異常數", "訪廠數", "結案數", "未結案數", "結案率(%)"],
        )
        detail_df = pd.DataFrame(
            [
                {
                    "日期": row["event_date"],
                    "類型": "異常" if row["event_type"] == "ANOMALY" else "訪廠",
                    "供應商": row["supplier_name"],
                    "問題/摘要": row["content"],
                    "狀態": row["status"],
                }
                for row in rows
            ]
        )
        output = Path(path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="月統計", index=False)
            ranking_df.to_excel(writer, sheet_name="供應商排行", index=False)
            detail_df.to_excel(writer, sheet_name="明細", index=False)
        return True, f"已匯出至：{output}"
    except Exception as exc:
        logger.exception("月報匯出失敗")
        return False, f"匯出失敗：{exc}"


def list_events_by_range(start_date: str, end_date: str) -> list[dict]:
    """取得指定日期範圍內的所有異常事件與訪廠事件。"""
    anomaly_sql = """
        SELECT
            a.id AS event_id,
            a.anomaly_no AS ref_no,
            a.anomaly_date AS event_date,
            'ANOMALY' AS event_type,
            s.supplier_name AS supplier_name,
            a.problem_desc AS content,
            a.status AS status,
            COALESCE(NULLIF(TRIM(a.root_cause_category), ''), a.category) AS category,
            a.root_cause_category AS root_cause_category,
            a.improvement_desc AS improvement_desc,
            a.closed_at AS closed_at
        FROM anomalies a
        JOIN suppliers s ON s.id = a.supplier_id
        WHERE a.anomaly_date BETWEEN ? AND ?
    """
    visit_sql = """
        SELECT
            v.id AS event_id,
            '' AS ref_no,
            v.visit_date AS event_date,
            'VISIT' AS event_type,
            s.supplier_name AS supplier_name,
            v.summary AS content,
            '已完成' AS status,
            '' AS category,
            '' AS root_cause_category,
            '' AS improvement_desc,
            NULL AS closed_at
        FROM visits v
        JOIN suppliers s ON s.id = v.supplier_id
        WHERE v.visit_date BETWEEN ? AND ?
    """
    events = []
    with get_connection() as conn:
        for row in conn.execute(anomaly_sql, (start_date, end_date)).fetchall():
            events.append(dict(row))
        for row in conn.execute(visit_sql, (start_date, end_date)).fetchall():
            events.append(dict(row))
    events.sort(key=lambda x: (x["event_date"], x["event_id"]), reverse=True)
    return events


def summarize_range_events(events: list[dict]) -> tuple[dict, list[dict]]:
    """依日期區間明細彙總 (整體 KPI, 供應商排行)。

    口徑：opened-in-range cohort 的「現況」統計（結案數 = 區間內發生且目前
    已結案），與 get_monthly_stats 的固定月份口徑（結案數 = 當月結案，跨
    cohort）刻意不同 — 兩者不可互相替換（audit finding D18：抽成具名純函
    式讓此口徑可獨立測試，而非埋在匯出流程內）。
    """
    from collections import defaultdict

    total_anomalies = len([e for e in events if e['event_type'] == 'ANOMALY'])
    total_visits = len([e for e in events if e['event_type'] == 'VISIT'])
    closed_anomalies = len([e for e in events if e['event_type'] == 'ANOMALY' and e['status'] == '已結案'])
    open_anomalies = len([e for e in events if e['event_type'] == 'ANOMALY' and e['status'] == '待處理'])
    totals = {
        "total_anomalies": total_anomalies,
        "total_visits": total_visits,
        "closed_anomalies": closed_anomalies,
        "open_anomalies": open_anomalies,
        "close_rate": (closed_anomalies / total_anomalies * 100) if total_anomalies > 0 else 0.0,
        "anomaly_visit_ratio": (total_anomalies / total_visits) if total_visits > 0 else 0.0,
        "supplier_coverage": len(set(e['supplier_name'] for e in events if e.get('supplier_name'))),
    }

    supplier_stats = defaultdict(lambda: {"anomaly_count": 0, "visit_count": 0, "closed_count": 0, "open_count": 0})
    for e in events:
        sname = e.get("supplier_name")
        if not sname:
            continue
        if e["event_type"] == "ANOMALY":
            supplier_stats[sname]["anomaly_count"] += 1
            if e["status"] == "已結案":
                supplier_stats[sname]["closed_count"] += 1
            else:
                supplier_stats[sname]["open_count"] += 1
        elif e["event_type"] == "VISIT":
            supplier_stats[sname]["visit_count"] += 1

    ranking_rows = []
    for sname, s in supplier_stats.items():
        tot_anom = s["anomaly_count"]
        cls_anom = s["closed_count"]
        rate = (cls_anom / tot_anom * 100) if tot_anom > 0 else 0.0
        ranking_rows.append({
            "supplier_name": sname,
            "anomaly_count": tot_anom,
            "visit_count": s["visit_count"],
            "closed_anomaly_count": cls_anom,
            "open_anomaly_count": s["open_count"],
            "close_rate_pct": rate
        })
    ranking_rows.sort(key=lambda x: x["anomaly_count"], reverse=True)
    return totals, ranking_rows


def _normalized_anomaly_category(value: object) -> str:
    text = str(value or "").strip()
    return text or "未分類"


def _build_category_pareto_rows(category_counts: dict[str, int]) -> list[dict]:
    total = sum(category_counts.values())
    if total <= 0:
        return []

    rows = []
    cumulative_count = 0
    sorted_items = sorted(
        category_counts.items(),
        key=lambda item: (-item[1], item[0] == "未分類", item[0]),
    )
    for rank, (category, count) in enumerate(sorted_items, start=1):
        cumulative_count += count
        rows.append({
            "rank": rank,
            "category": category,
            "count": int(count),
            "percent": round(count / total * 100, 1),
            "cumulative_percent": round(cumulative_count / total * 100, 1),
        })
    rows[-1]["cumulative_percent"] = 100.0
    return rows


def get_anomaly_category_pareto_by_range(start_date: str, end_date: str) -> list[dict]:
    """Return root-cause Pareto rows for anomalies opened in a date range.

    頁面圖表與 Excel 匯出(表格、嵌入 PNG)都必須走這個唯一實作,
    避免兩套彙總口徑(fallback、JOIN 掉列)造成同一份報告內數字不一致。
    """
    from collections import defaultdict

    sql = """
        SELECT
            COALESCE(
                NULLIF(TRIM(root_cause_category), ''),
                NULLIF(TRIM(category), ''),
                '未分類'
            ) AS category,
            COUNT(*) AS count
        FROM anomalies
        WHERE anomaly_date BETWEEN ? AND ?
        GROUP BY COALESCE(
            NULLIF(TRIM(root_cause_category), ''),
            NULLIF(TRIM(category), ''),
            '未分類'
        )
        ORDER BY count DESC, category ASC
    """
    with get_connection() as conn:
        rows = conn.execute(sql, (start_date, end_date)).fetchall()
    # SQL TRIM 只去 ASCII 空白,Python strip() 另涵蓋全形空白等 Unicode 空白;
    # 兩個 SQL 群組正規化成同一鍵時必須累加,不能讓後者覆蓋前者的計數。
    category_counts: dict[str, int] = defaultdict(int)
    for row in rows:
        category_counts[_normalized_anomaly_category(row["category"])] += int(row["count"] or 0)
    return _build_category_pareto_rows(dict(category_counts))


def export_events_report(
    file_path: str,
    start_date: str,
    end_date: str,
    temp_chart_paths: dict[str, str] | None = None,
) -> tuple[bool, str]:
    """匯出格式優化後的供應商異常事件統計分析報告，包含視覺總覽與明細/排行表格。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from datetime import datetime

    try:
        events = list_events_by_range(start_date, end_date)

        totals, ranking_rows = summarize_range_events(events)
        category_pareto_rows = get_anomaly_category_pareto_by_range(start_date, end_date)
        total_anomalies = totals["total_anomalies"]
        total_visits = totals["total_visits"]
        closed_anomalies = totals["closed_anomalies"]
        open_anomalies = totals["open_anomalies"]
        close_rate = totals["close_rate"]
        anomaly_visit_ratio = totals["anomaly_visit_ratio"]
        supplier_coverage = totals["supplier_coverage"]

        workbook = Workbook()
        
        # 樣式定義
        FONT_NAME = "Microsoft JhengHei"
        STYLE_FONT = Font(name=FONT_NAME, size=11)
        STYLE_FONT_BOLD = Font(name=FONT_NAME, size=11, bold=True)
        STYLE_HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        STYLE_TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color="1E3A8A")
        STYLE_SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")

        STYLE_FILL_HEADER = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        STYLE_FILL_ZEBRA = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        STYLE_FILL_KPI_BG = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        STYLE_FILL_TOTAL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        STYLE_BORDER_THIN = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )
        STYLE_BORDER_TOTAL = Border(
            top=Side(style="thin", color="9CA3AF"),
            bottom=Side(style="double", color="111827")
        )

        ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
        ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

        def _auto_fit(ws):
            for col in ws.columns:
                vals = [str(c.value or "") for c in col]
                max_len = max((len(v.encode('utf-8')) for v in vals), default=0)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

        # 1. 視覺總覽報告頁
        report_sheet = workbook.active
        report_sheet.title = "統計報告"
        report_sheet.views.sheetView[0].showGridLines = True

        report_sheet.cell(row=1, column=1, value="供應商品質異常事件統計分析報告").font = STYLE_TITLE_FONT
        report_sheet.row_dimensions[1].height = 30
        
        subtitle_text = f"統計區間：{start_date} 至 {end_date}   |   報告生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        report_sheet.cell(row=2, column=1, value=subtitle_text).font = STYLE_SUBTITLE_FONT

        # KPI 卡片一 (A4:C5)
        report_sheet.merge_cells("A4:C4")
        report_sheet.cell(row=4, column=1, value="📊 異常事件統計總覽").font = STYLE_FONT_BOLD
        report_sheet.cell(row=4, column=1).alignment = ALIGN_CENTER
        report_sheet.cell(row=4, column=1).fill = STYLE_FILL_KPI_BG
        
        report_sheet.cell(row=5, column=1, value=f"總異常件數: {total_anomalies} 件").font = STYLE_FONT
        report_sheet.cell(row=5, column=1).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=2, value=f"總訪廠件數: {total_visits} 件").font = STYLE_FONT
        report_sheet.cell(row=5, column=2).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=3, value=f"異常/訪廠比: {anomaly_visit_ratio:.2f}").font = STYLE_FONT
        report_sheet.cell(row=5, column=3).alignment = ALIGN_CENTER

        # KPI 卡片二 (E4:G5)
        report_sheet.merge_cells("E4:G4")
        report_sheet.cell(row=4, column=5, value="🎯 處理績效與涵蓋率").font = STYLE_FONT_BOLD
        report_sheet.cell(row=4, column=5).alignment = ALIGN_CENTER
        report_sheet.cell(row=4, column=5).fill = STYLE_FILL_KPI_BG
        
        report_sheet.cell(row=5, column=5, value=f"已結案/未結案: {closed_anomalies} / {open_anomalies}").font = STYLE_FONT
        report_sheet.cell(row=5, column=5).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=6, value=f"結案率: {close_rate:.1f}%").font = STYLE_FONT
        report_sheet.cell(row=5, column=6).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=7, value=f"供應商覆蓋數: {supplier_coverage} 家").font = STYLE_FONT
        report_sheet.cell(row=5, column=7).alignment = ALIGN_CENTER

        for r in [4, 5]:
            for c in [1, 2, 3, 5, 6, 7]:
                report_sheet.cell(row=r, column=c).border = STYLE_BORDER_THIN

        # 插入統計圖表 (橫向與縱向並排)
        if temp_chart_paths:
            chart_placements = [
                ("trend", "A7"),
                ("visit_anomaly", "I7"),
                ("responsible", "A23"),
                ("category_pareto", "I23"),
            ]
            for key, cell in chart_placements:
                path = temp_chart_paths.get(key)
                if path and Path(path).exists():
                    img = Image(path)
                    img.width = 460
                    img.height = 310
                    report_sheet.add_image(img, cell)

        # 2. 異常類別柏拉圖資料頁
        category_sheet = workbook.create_sheet("異常類別柏拉圖")
        category_sheet.views.sheetView[0].showGridLines = True

        category_headers = ["排名", "異常類別", "件數", "佔比(%)", "累積佔比(%)"]
        category_sheet.append(category_headers)
        for col_idx in range(1, len(category_headers) + 1):
            cell = category_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        category_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(category_pareto_rows, start=2):
            data = [
                row.get("rank", 0),
                row.get("category", ""),
                row.get("count", 0),
                row.get("percent", 0.0),
                row.get("cumulative_percent", 0.0),
            ]
            category_sheet.append(data)

            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(category_headers) + 1):
                cell = category_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA
                if c_idx == 2:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
                if c_idx in (4, 5):
                    cell.number_format = "0.0"
            category_sheet.row_dimensions[r_idx].height = 20
        _auto_fit(category_sheet)

        # 3. 異常事件明細頁
        detail_sheet = workbook.create_sheet("異常事件明細")
        detail_sheet.views.sheetView[0].showGridLines = True
        
        headers = ["日期", "類型", "供應商名稱", "問題與摘要說明", "當前狀態", "類別", "改善說明", "結案日期"]
        detail_sheet.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = detail_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        detail_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(events, start=2):
            data = [
                row.get("event_date", ""),
                "異常" if row.get("event_type") == "ANOMALY" else "訪廠",
                row.get("supplier_name", ""),
                row.get("content", ""),
                row.get("status", ""),
                row.get("category", ""),
                row.get("improvement_desc", ""),
                row.get("closed_at", "")
            ]
            detail_sheet.append(data)
            
            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(headers) + 1):
                cell = detail_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA
                
                # 對齊方式
                if c_idx in (1, 2, 5, 8):
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = ALIGN_LEFT
            detail_sheet.row_dimensions[r_idx].height = 20
        _auto_fit(detail_sheet)

        # 4. 供應商排行榜頁
        rank_sheet = workbook.create_sheet("供應商排行榜")
        rank_sheet.views.sheetView[0].showGridLines = True
        
        rank_headers = ["排名", "供應商名稱", "異常事件數", "訪廠次數", "已結案數", "未結案數", "結案率(%)"]
        rank_sheet.append(rank_headers)
        for col_idx in range(1, len(rank_headers) + 1):
            cell = rank_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        rank_sheet.row_dimensions[1].height = 24

        for r_idx, row in enumerate(ranking_rows, start=2):
            data = [
                r_idx - 1,
                row.get("supplier_name", ""),
                row.get("anomaly_count", 0),
                row.get("visit_count", 0),
                row.get("closed_anomaly_count", 0),
                row.get("open_anomaly_count", 0),
                f"{row.get('close_rate_pct', 0.0):.1f}%"
            ]
            rank_sheet.append(data)
            
            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(rank_headers) + 1):
                cell = rank_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA
                
                if c_idx == 1:
                    cell.alignment = ALIGN_CENTER
                elif c_idx == 2:
                    cell.alignment = ALIGN_LEFT
                else:
                    cell.alignment = ALIGN_RIGHT
            rank_sheet.row_dimensions[r_idx].height = 20

        # 合計列
        total_row_idx = len(ranking_rows) + 2
        if len(ranking_rows) > 0:
            rank_sheet.cell(row=total_row_idx, column=1, value="合計").font = STYLE_FONT_BOLD
            rank_sheet.cell(row=total_row_idx, column=1).alignment = ALIGN_CENTER
            
            for c_idx in (3, 4, 5, 6):
                col_letter = get_column_letter(c_idx)
                sum_formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                cell = rank_sheet.cell(row=total_row_idx, column=c_idx, value=sum_formula)
                cell.font = STYLE_FONT_BOLD
                cell.alignment = ALIGN_RIGHT
                cell.number_format = "#,##0"

            # 總體結案率公式 = 總已結案數 / 總異常件數
            total_closed_cell = f"E{total_row_idx}"
            total_anomaly_cell = f"C{total_row_idx}"
            rate_formula = f"=IF({total_anomaly_cell}>0, {total_closed_cell}/{total_anomaly_cell}, 0)"
            rate_cell = rank_sheet.cell(row=total_row_idx, column=7, value=rate_formula)
            rate_cell.font = STYLE_FONT_BOLD
            rate_cell.alignment = ALIGN_RIGHT
            rate_cell.number_format = "0.0%"

            for c_idx in range(1, len(rank_headers) + 1):
                cell = rank_sheet.cell(row=total_row_idx, column=c_idx)
                cell.border = STYLE_BORDER_TOTAL
                cell.fill = STYLE_FILL_TOTAL

        _auto_fit(rank_sheet)

        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return True, f"已匯出至：{output_path}"
    except Exception as exc:
        logger.exception("自訂日期區間 Excel 報告匯出出錯")
        return False, f"匯出報告失敗：{exc}"


def get_responsible_person_stats_by_range(start_date: str, end_date: str) -> list[dict]:
    """計算指定日期範圍內各責任人的異常件數與平均處理時效。"""
    sql = """
        SELECT 
            COALESCE(NULLIF(TRIM(responsible_person), ''), '未指定') AS person,
            COUNT(*) AS total_count,
            COUNT(CASE WHEN status = '已結案' THEN 1 END) AS closed_count,
            COUNT(CASE WHEN status = '待處理' THEN 1 END) AS open_count,
            AVG(julianday(COALESCE(NULLIF(closed_at, ''), date('now', 'localtime'))) - julianday(anomaly_date)) AS avg_days
        FROM anomalies
        WHERE anomaly_date BETWEEN ? AND ?
        GROUP BY person
        ORDER BY total_count DESC, person ASC
    """
    with get_connection() as conn:
        rows = conn.execute(sql, (start_date, end_date)).fetchall()
        
        # 未結案最早/最晚日期必須與 open_count 同一區間口徑(圖表長條、tooltip、
        # 洞察文字皆以區間內事件為準),否則日期會指向不在計數內的案件。
        unclosed_sql = """
            SELECT
                COALESCE(NULLIF(TRIM(responsible_person), ''), '未指定') AS person,
                MIN(anomaly_date) AS min_date,
                MAX(anomaly_date) AS max_date
            FROM anomalies
            WHERE status = '待處理' AND anomaly_date BETWEEN ? AND ?
            GROUP BY person
        """
        unclosed_rows = conn.execute(unclosed_sql, (start_date, end_date)).fetchall()
        unclosed_dates = {r["person"]: (r["min_date"], r["max_date"]) for r in unclosed_rows}
    
    results = []
    for row in rows:
        person = row["person"]
        min_date, max_date = unclosed_dates.get(person, (None, None))
        results.append({
            "responsible_person": person,
            "total_count": int(row["total_count"]),
            "closed_count": int(row["closed_count"]),
            "open_count": int(row["open_count"]),
            "avg_resolution_time": round(float(row["avg_days"] or 0), 1),
            "min_open_date": min_date,
            "max_open_date": max_date,
        })
    return results


def get_visit_trend_by_range(start_date: str, end_date: str) -> list[dict]:
    """計算指定日期範圍內各月份的訪廠數與訪廠發現的異常數（最多限制 12 個月）。"""
    from datetime import datetime
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    except Exception:
        return []
    
    months_list = []
    curr_y, curr_m = start_dt.year, start_dt.month
    end_y, end_m = end_dt.year, end_dt.month
    
    while (curr_y < end_y) or (curr_y == end_y and curr_m <= end_m):
        months_list.append(f"{curr_y:04d}-{curr_m:02d}")
        curr_m += 1
        if curr_m > 12:
            curr_m = 1
            curr_y += 1
            
    if len(months_list) > 12:
        months_list = months_list[-12:]
    elif not months_list:
        months_list = [start_date[:7]]

    with get_connection() as conn:
        visit_rows = conn.execute(
            """
            SELECT substr(visit_date, 1, 7) AS yyyymm, COUNT(*) AS visit_count
            FROM visits
            WHERE visit_date BETWEEN ? AND ?
            GROUP BY yyyymm
            """,
            (start_date, end_date)
        ).fetchall()
        visits_by_month = {r["yyyymm"]: int(r["visit_count"] or 0) for r in visit_rows}

        anomaly_rows = conn.execute(
            """
            SELECT substr(anomaly_date, 1, 7) AS yyyymm, COUNT(*) AS anomaly_count
            FROM anomalies
            WHERE NULLIF(visit_id, '') IS NOT NULL AND anomaly_date BETWEEN ? AND ?
            GROUP BY yyyymm
            """,
            (start_date, end_date)
        ).fetchall()
        anomalies_by_month = {r["yyyymm"]: int(r["anomaly_count"] or 0) for r in anomaly_rows}

    results = []
    for yyyymm in months_list:
        results.append({
            "yyyymm": yyyymm,
            "visit_count": visits_by_month.get(yyyymm, 0),
            "visit_anomaly_count": anomalies_by_month.get(yyyymm, 0)
        })
    return results


def get_anomaly_trend_by_range(start_date: str, end_date: str) -> list[dict]:
    """計算指定日期範圍內各月份的異常數、結案數、逾期數及累計積壓趨勢（最多限制 12 個月）。"""
    import calendar
    from datetime import datetime
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    except Exception:
        return []
    
    months_list = []
    curr_y, curr_m = start_dt.year, start_dt.month
    end_y, end_m = end_dt.year, end_dt.month
    
    while (curr_y < end_y) or (curr_y == end_y and curr_m <= end_m):
        months_list.append(f"{curr_y:04d}-{curr_m:02d}")
        curr_m += 1
        if curr_m > 12:
            curr_m = 1
            curr_y += 1
            
    if len(months_list) > 12:
        months_list = months_list[-12:]
    elif not months_list:
        months_list = [start_date[:7]]

    def _month_end_date(yyyymm: str) -> str:
        year, month = int(yyyymm[:4]), int(yyyymm[5:])
        last_day = calendar.monthrange(year, month)[1]
        return f"{year:04d}-{month:02d}-{last_day:02d}"

    with get_connection() as conn:
        # total + overdue per month in one grouped query (both keyed off
        # anomaly_date; overdue's "now" cutoff doesn't depend on the month
        # being reported, so it groups cleanly alongside total). Keep the
        # date predicate exact so custom export ranges do not include
        # same-month rows outside the selected start/end dates.
        total_overdue_rows = conn.execute(
            """
            SELECT
                substr(anomaly_date, 1, 7) AS yyyymm,
                COUNT(*) AS total_count,
                SUM(
                    CASE WHEN status = '待處理' AND due_date <> ''
                              AND due_date < date('now', 'localtime')
                         THEN 1 ELSE 0 END
                ) AS overdue_count
            FROM anomalies
            WHERE anomaly_date BETWEEN ? AND ?
            GROUP BY yyyymm
            """,
            (start_date, end_date),
        ).fetchall()
        total_by_month = {r["yyyymm"]: int(r["total_count"] or 0) for r in total_overdue_rows}
        overdue_by_month = {r["yyyymm"]: int(r["overdue_count"] or 0) for r in total_overdue_rows}

        # closed per month, grouped by closed_at (a different column than
        # anomaly_date, so it needs its own query).
        closed_rows = conn.execute(
            """
            SELECT substr(closed_at, 1, 7) AS yyyymm, COUNT(*) AS closed_count
            FROM anomalies
            WHERE closed_at <> '' AND closed_at BETWEEN ? AND ?
            GROUP BY yyyymm
            """,
            (start_date, end_date),
        ).fetchall()
        closed_by_month = {r["yyyymm"]: int(r["closed_count"] or 0) for r in closed_rows}

        # backlog: whether a row counts as "still open as of yyyymm" depends
        # on yyyymm itself (not a fixed cutoff), so it can't be grouped in
        # one pass -- fetch every candidate row once and re-apply the
        # original per-month condition in Python instead of re-scanning the
        # table once per month (was up to 12 additional full-table scans).
        all_rows = conn.execute(
            "SELECT anomaly_date, status, closed_at FROM anomalies "
            "WHERE anomaly_date <= ?",
            (end_date,),
        ).fetchall()

    results = []
    for yyyymm in months_list:
        cutoff_date = min(_month_end_date(yyyymm), end_date)
        backlog_count = 0
        for row in all_rows:
            row_date = str(row["anomaly_date"] or "")
            if row_date > cutoff_date:
                continue
            status = row["status"]
            closed_at = row["closed_at"] or ""
            is_open_as_of_month = (
                status != "已結案"
                or (closed_at != "" and closed_at > cutoff_date)
            )
            if is_open_as_of_month:
                backlog_count += 1
        results.append({
            "yyyymm": yyyymm,
            "total_count": total_by_month.get(yyyymm, 0),
            "closed_count": closed_by_month.get(yyyymm, 0),
            "overdue_count": overdue_by_month.get(yyyymm, 0),
            "backlog_count": backlog_count,
        })
    return results

