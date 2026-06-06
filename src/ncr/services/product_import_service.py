from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ncr.db import crud


STATUS_CAN_ADD = "可新增"
STATUS_CAN_UPDATE = "可更新"
STATUS_SKIPPED = "已一致略過"
STATUS_BLOCKED = "錯誤擋下"
ITEM_NO_HEADERS = {"料號", "產品料號", "item_no"}
PRODUCT_NAME_HEADERS = {"產品名稱", "product_name"}


class ProductImportError(ValueError):
    """Raised when an Excel file cannot be opened or parsed."""


@dataclass(frozen=True)
class ProductImportRow:
    row_number: int
    status: str
    item_no: str
    product_name: str
    current_product_name: str
    message: str
    defect_update_count: int = 0
    will_update_product: bool = False

    @property
    def can_add(self) -> bool:
        return self.status == STATUS_CAN_ADD

    @property
    def can_update(self) -> bool:
        return self.status == STATUS_CAN_UPDATE

    @property
    def is_error(self) -> bool:
        return self.status == STATUS_BLOCKED


@dataclass(frozen=True)
class ProductImportPreview:
    rows: list[ProductImportRow]
    file_errors: list[str]

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    @property
    def add_count(self) -> int:
        return sum(1 for row in self.rows if row.can_add)

    @property
    def importable_count(self) -> int:
        return self.add_count

    @property
    def product_update_count(self) -> int:
        return sum(1 for row in self.rows if row.will_update_product)

    @property
    def defect_update_count(self) -> int:
        return sum(row.defect_update_count for row in self.rows)

    @property
    def skipped_count(self) -> int:
        return sum(1 for row in self.rows if row.status == STATUS_SKIPPED)

    @property
    def error_count(self) -> int:
        row_errors = sum(1 for row in self.rows if row.is_error)
        return len(self.file_errors) + row_errors

    @property
    def can_import(self) -> bool:
        return self.error_count == 0

    @property
    def has_writes(self) -> bool:
        return (
            self.add_count > 0
            or self.product_update_count > 0
            or self.defect_update_count > 0
        )

    def rows_to_import(self) -> list[ProductImportRow]:
        return [row for row in self.rows if row.can_add]

    def rows_to_product_update(self) -> list[ProductImportRow]:
        return [row for row in self.rows if row.will_update_product]

    def rows_to_defect_update(self) -> list[ProductImportRow]:
        return [row for row in self.rows if row.defect_update_count > 0]


@dataclass(frozen=True)
class ProductImportResult:
    added_count: int
    product_updated_count: int
    defect_updated_count: int
    skipped_count: int
    backup_path: Path | None


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return _normalize_cell(value).lower()


def _find_column_index(headers: list[Any], accepted_names: set[str]) -> int | None:
    accepted = {name.lower() for name in accepted_names}
    for index, header in enumerate(headers):
        if _normalize_header(header) in accepted:
            return index
    return None


def _read_workbook_rows(
    file_path: str | Path,
) -> tuple[list[Any], list[tuple[int, list[Any]]]]:
    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise ProductImportError("僅支援 .xlsx 檔案。")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, BadZipFile, InvalidFileException) as exc:
        raise ProductImportError(f"無法開啟 Excel 檔案：{exc}") from exc
    except ValueError as exc:
        raise ProductImportError(f"Excel 檔案格式無法讀取：{exc}") from exc

    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None or not any(_normalize_cell(cell) for cell in header_row):
            return [], []
        data_rows: list[tuple[int, list[Any]]] = []
        for row_number, row_values in enumerate(rows_iter, start=2):
            values = list(row_values)
            if any(_normalize_cell(cell) for cell in values):
                data_rows.append((row_number, values))
        return list(header_row), data_rows
    finally:
        workbook.close()


def preview_product_import(
    conn: sqlite3.Connection, file_path: str | Path
) -> ProductImportPreview:
    headers, data_rows = _read_workbook_rows(file_path)
    file_errors: list[str] = []
    preview_rows: list[ProductImportRow] = []
    if not headers:
        return ProductImportPreview(rows=[], file_errors=["Excel 第一列需包含欄位標題。"])

    item_no_index = _find_column_index(headers, ITEM_NO_HEADERS)
    product_name_index = _find_column_index(headers, PRODUCT_NAME_HEADERS)
    if item_no_index is None:
        file_errors.append("找不到料號欄位，請使用「料號」、「產品料號」或「item_no」。")
    if product_name_index is None:
        file_errors.append("找不到產品名稱欄位，請使用「產品名稱」或「product_name」。")
    if file_errors:
        return ProductImportPreview(rows=[], file_errors=file_errors)
    if not data_rows:
        return ProductImportPreview(rows=[], file_errors=["Excel 沒有可匯入的資料列。"])

    normalized_rows: list[tuple[int, str, str]] = []
    item_no_counts: dict[str, int] = {}
    for row_number, values in data_rows:
        item_no = _normalize_cell(values[item_no_index]) if item_no_index < len(values) else ""
        product_name = (
            _normalize_cell(values[product_name_index])
            if product_name_index < len(values)
            else ""
        )
        normalized_rows.append((row_number, item_no, product_name))
        if item_no:
            item_no_counts[item_no] = item_no_counts.get(item_no, 0) + 1

    for row_number, item_no, product_name in normalized_rows:
        if not item_no:
            preview_rows.append(
                ProductImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    "",
                    product_name,
                    "",
                    "料號不可空白。",
                )
            )
            continue
        if not product_name:
            preview_rows.append(
                ProductImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    item_no,
                    "",
                    "",
                    "產品名稱不可空白。",
                )
            )
            continue
        if item_no_counts.get(item_no, 0) > 1:
            preview_rows.append(
                ProductImportRow(
                    row_number,
                    STATUS_BLOCKED,
                    item_no,
                    product_name,
                    "",
                    "Excel 內部有重複料號，請先整理來源資料。",
                )
            )
            continue

        defect_update_count = crud.count_defect_product_name_changes(
            conn,
            item_no,
            product_name,
        )
        existing = crud.get_product_by_item_no(conn, item_no)
        if existing is None:
            message = "將新增至產品清單。"
            if defect_update_count:
                message = f"{message} 歷史不良品將修正 {defect_update_count} 筆。"
            preview_rows.append(
                ProductImportRow(
                    row_number,
                    STATUS_CAN_ADD,
                    item_no,
                    product_name,
                    "",
                    message,
                    defect_update_count,
                )
            )
            continue
        current_product_name = _normalize_cell(existing["product_name"])
        if current_product_name == product_name:
            if defect_update_count:
                preview_rows.append(
                    ProductImportRow(
                        row_number,
                        STATUS_CAN_UPDATE,
                        item_no,
                        product_name,
                        current_product_name,
                        f"產品清單已一致；歷史不良品將修正 {defect_update_count} 筆。",
                        defect_update_count,
                    )
                )
            else:
                preview_rows.append(
                    ProductImportRow(
                        row_number,
                        STATUS_SKIPPED,
                        item_no,
                        product_name,
                        current_product_name,
                        "產品清單與歷史不良品已一致，匯入時略過。",
                    )
                )
        else:
            preview_rows.append(
                ProductImportRow(
                    row_number,
                    STATUS_CAN_UPDATE,
                    item_no,
                    product_name,
                    current_product_name,
                    (
                        "將以 Excel 產品名稱更新產品清單；"
                        f"歷史不良品將修正 {defect_update_count} 筆。"
                    ),
                    defect_update_count,
                    True,
                )
            )
    return ProductImportPreview(rows=preview_rows, file_errors=file_errors)


def _database_file_path(conn: sqlite3.Connection) -> Path | None:
    for row in conn.execute("PRAGMA database_list").fetchall():
        name = row["name"] if isinstance(row, sqlite3.Row) else row[1]
        file_name = row["file"] if isinstance(row, sqlite3.Row) else row[2]
        if name != "main" or not file_name or file_name == ":memory:":
            continue
        path = Path(str(file_name))
        if path.exists():
            return path
    return None


def _create_database_backup(conn: sqlite3.Connection) -> Path | None:
    source_path = _database_file_path(conn)
    if source_path is None:
        return None
    backup_dir = source_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"defect-before-product-import-{timestamp}.db"
    suffix = 1
    while backup_path.exists():
        backup_path = backup_dir / (
            f"defect-before-product-import-{timestamp}-{suffix}.db"
        )
        suffix += 1
    backup_conn = sqlite3.connect(backup_path)
    try:
        conn.backup(backup_conn)
    finally:
        backup_conn.close()
    return backup_path


def _product_payloads(
    rows: list[ProductImportRow],
    created_at: str,
) -> list[dict[str, str]]:
    return [
        {
            "item_no": row.item_no,
            "product_name": row.product_name,
            "created_at": created_at,
        }
        for row in rows
    ]


def apply_product_import(
    conn: sqlite3.Connection, preview: ProductImportPreview
) -> ProductImportResult:
    if not preview.can_import:
        raise ValueError("匯入清單仍有錯誤，請修正 Excel 後重新預覽。")
    if not preview.has_writes:
        return ProductImportResult(
            added_count=0,
            product_updated_count=0,
            defect_updated_count=0,
            skipped_count=preview.skipped_count,
            backup_path=None,
        )
    now = datetime.now().isoformat(timespec="seconds")
    import_rows = preview.rows_to_import()
    product_update_rows = preview.rows_to_product_update()
    defect_update_rows = preview.rows_to_defect_update()
    added_payloads = _product_payloads(import_rows, now)
    product_update_payloads = _product_payloads(product_update_rows, now)
    defect_update_payloads = _product_payloads(defect_update_rows, now)
    backup_path = _create_database_backup(conn)
    conn.execute("BEGIN")
    try:
        added_count = crud.insert_products(conn, added_payloads)
        product_updated_count = crud.update_products_by_item_no(
            conn,
            product_update_payloads,
        )
        defect_updated_count = crud.update_defect_product_names_by_item_no(
            conn,
            defect_update_payloads,
        )
        conn.commit()
        return ProductImportResult(
            added_count=added_count,
            product_updated_count=product_updated_count,
            defect_updated_count=defect_updated_count,
            skipped_count=preview.skipped_count,
            backup_path=backup_path,
        )
    except sqlite3.Error:
        conn.rollback()
        raise
