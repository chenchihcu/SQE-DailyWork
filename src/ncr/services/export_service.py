from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from ncr.models.defect import (
    DETAIL_EXPORT_COLUMNS,
    DETAIL_EXPORT_HEADERS,
    STATS_DIMENSION_COLUMNS,
    STATS_SECTION_DEFINITIONS,
    build_stats_headers,
)

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parents[3] / "Outputs"

# ── 樣式定義 ────────────────────────────────────────────────────────
FONT_NAME = "Microsoft JhengHei"
STYLE_FONT = Font(name=FONT_NAME, size=11)
STYLE_FONT_BOLD = Font(name=FONT_NAME, size=11, bold=True)
STYLE_HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
STYLE_TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color="1E3A8A")
STYLE_SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")

STYLE_FILL_HEADER = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
STYLE_FILL_ZEBRA = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
STYLE_FILL_KPI_BG = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
STYLE_FILL_TOTAL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

STYLE_BORDER_THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB")
)
STYLE_BORDER_TOTAL = Border(
    top=Side(style="thin", color="9CA3AF"),
    bottom=Side(style="double", color="111827")
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

def _normalize_rows(rows: Iterable) -> list[dict]:
    normalized: list[dict] = []
    for row in rows:
        if isinstance(row, dict):
            normalized.append(row)
        else:
            normalized.append(dict(row))
    return normalized


def _summarize_rows(rows: list[dict], key_name: str, *, skip_blank: bool) -> list[dict]:
    summary: dict[tuple[str, str, str, str, str], tuple[int, int]] = {}
    for row in rows:
        name = str(row.get(key_name, "") or "").strip()
        if skip_blank and (not name or name == "N/A"):
            continue
        disposition = str(row.get("disposition", "") or "").strip()
        category = str(row.get("category", "") or "").strip()
        status = str(row.get("status", "") or "").strip()
        event_date = str(row.get("event_date", "") or "").strip()
        event_month = event_date[:7] if len(event_date) >= 7 else ""
        try:
            qty_value = int(row.get("qty", 0) or 0)
        except (TypeError, ValueError):
            qty_value = 0

        group_key = (name, disposition, category, event_month, status)
        existing = summary.setdefault(group_key, (0, 0))
        case_count, total_qty = existing
        summary[group_key] = (case_count + 1, total_qty + qty_value)

    def _month_sort_token(value: str) -> int:
        token = value.replace("-", "")
        return int(token) if token.isdigit() else 0

    sorted_rows = sorted(
        summary.items(),
        key=lambda item: (
            -item[1][1],
            item[0][0],
            -_month_sort_token(item[0][3]),
            item[0][1],
            item[0][2],
            item[0][4],
        ),
    )
    return [
        {
            key_name: name,
            "disposition": disposition,
            "category": category,
            "event_month": event_month,
            "status": status,
            "case_count": case_count,
            "total_qty": total_qty,
        }
        for (
            name,
            disposition,
            category,
            event_month,
            status,
        ), (case_count, total_qty) in sorted_rows
    ]


def _auto_fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max((len(value.encode('utf-8')) for value in values), default=0)
        # 用 utf-8 byte 長度來適配中文字寬
        col_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 45)


def export_ncr_excel_report(
    file_path: str,
    start_date: str,
    end_date: str,
    defects: list[dict],
    temp_chart_paths: dict[str, str] | None = None,
) -> tuple[bool, str]:
    """匯出格式優化後的倉庫不合格品統計報告，包含視覺總覽與明細/統計表格。"""
    try:
        defect_rows = _normalize_rows(defects)
        product_rows = _summarize_rows(defect_rows, "product_name", skip_blank=False)
        supplier_rows = _summarize_rows(defect_rows, "supplier_name", skip_blank=True)
        outsource_rows = _summarize_rows(defect_rows, "outsource_supplier_name", skip_blank=True)

        workbook = Workbook()
        
        # 1. 視覺總覽報告頁 (統計報告)
        report_sheet = workbook.active
        report_sheet.title = "統計報告"
        report_sheet.views.sheetView[0].showGridLines = True

        # 報告標題與時間資訊
        report_sheet.cell(row=1, column=1, value="倉庫不合格品統計分析報告").font = STYLE_TITLE_FONT
        report_sheet.row_dimensions[1].height = 30
        
        subtitle_text = f"統計區間：{start_date} 至 {end_date}   |   報告生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        report_sheet.cell(row=2, column=1, value=subtitle_text).font = STYLE_SUBTITLE_FONT
        
        # 核心 KPI 卡片 (A4:C5)
        total_cases = len(defect_rows)
        total_qty = sum(int(row.get("qty") or 0) for row in defect_rows)
        scrap_qty = sum(int(row.get("qty") or 0) for row in defect_rows if row.get("disposition") == "報廢")
        rework_qty = sum(int(row.get("qty") or 0) for row in defect_rows if row.get("disposition") == "重工")
        scrap_rate = (scrap_qty / total_qty * 100) if total_qty > 0 else 0.0

        # 卡片一：總量卡片
        report_sheet.merge_cells("A4:C4")
        report_sheet.cell(row=4, column=1, value="📊 不合格品總量統計").font = STYLE_FONT_BOLD
        report_sheet.cell(row=4, column=1).alignment = ALIGN_CENTER
        report_sheet.cell(row=4, column=1).fill = STYLE_FILL_KPI_BG
        
        report_sheet.cell(row=5, column=1, value=f"總件數: {total_cases} 件").font = STYLE_FONT
        report_sheet.cell(row=5, column=1).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=2, value=f"總數量: {total_qty} 件").font = STYLE_FONT
        report_sheet.cell(row=5, column=2).alignment = ALIGN_CENTER
        
        # 卡片二：處置卡片
        report_sheet.merge_cells("E4:G4")
        report_sheet.cell(row=4, column=5, value="⚙️ 處置方式與損失率").font = STYLE_FONT_BOLD
        report_sheet.cell(row=4, column=5).alignment = ALIGN_CENTER
        report_sheet.cell(row=4, column=5).fill = STYLE_FILL_KPI_BG
        
        report_sheet.cell(row=5, column=5, value=f"報廢數量: {scrap_qty} 件 ({scrap_rate:.1f}%)").font = STYLE_FONT
        report_sheet.cell(row=5, column=5).alignment = ALIGN_CENTER
        report_sheet.cell(row=5, column=6, value=f"重工數量: {rework_qty} 件").font = STYLE_FONT
        report_sheet.cell(row=5, column=6).alignment = ALIGN_CENTER

        # 套用卡片框線
        for r in [4, 5]:
            for c in [1, 2, 3, 5, 6, 7]:
                report_sheet.cell(row=r, column=c).border = STYLE_BORDER_THIN

        # 插入統計圖表 (2x2 網格)
        if temp_chart_paths:
            chart_placements = [
                ("supplier", "A7"),
                ("product", "I7"),
                ("disposition", "A26"),
                ("return_slip", "I26")
            ]
            for key, cell in chart_placements:
                path = temp_chart_paths.get(key)
                if path and Path(path).exists():
                    img = Image(path)
                    # 設定 Excel 圖表合適的尺寸比例
                    img.width = 460
                    img.height = 310
                    report_sheet.add_image(img, cell)

        # 2. 不良品明細頁
        detail_sheet = workbook.create_sheet("不良品明細")
        detail_sheet.views.sheetView[0].showGridLines = True
        
        # 表頭
        detail_sheet.append(DETAIL_EXPORT_HEADERS)
        for col_idx in range(1, len(DETAIL_EXPORT_HEADERS) + 1):
            cell = detail_sheet.cell(row=1, column=col_idx)
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = STYLE_BORDER_THIN
        detail_sheet.row_dimensions[1].height = 24

        # 數據行
        start_row = 2
        for r_idx, row in enumerate(defect_rows, start=start_row):
            data = [row.get(field_name, "") for field_name, _ in DETAIL_EXPORT_COLUMNS]
            detail_sheet.append(data)
            
            # 套用斑馬紋與字體框線
            is_even = (r_idx % 2 == 0)
            for c_idx in range(1, len(DETAIL_EXPORT_HEADERS) + 1):
                cell = detail_sheet.cell(row=r_idx, column=c_idx)
                cell.font = STYLE_FONT
                cell.border = STYLE_BORDER_THIN
                if is_even:
                    cell.fill = STYLE_FILL_ZEBRA
                
                # 特定欄位對齊方式
                field_name = DETAIL_EXPORT_COLUMNS[c_idx - 1][0]
                if field_name == "qty":
                    cell.alignment = ALIGN_RIGHT
                    cell.number_format = "#,##0"
                elif field_name in ("defect_no", "event_date", "status", "disposition", "category"):
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = ALIGN_LEFT
            detail_sheet.row_dimensions[r_idx].height = 20

        # 合計行
        total_row_idx = len(defect_rows) + start_row
        if len(defect_rows) > 0:
            detail_sheet.cell(row=total_row_idx, column=1, value="合計").font = STYLE_FONT_BOLD
            detail_sheet.cell(row=total_row_idx, column=1).alignment = ALIGN_CENTER
            
            # 在數量欄位寫入 SUM 公式 (第九欄為 qty)
            qty_col_letter = get_column_letter(9)
            sum_formula = f"=SUM({qty_col_letter}2:{qty_col_letter}{total_row_idx - 1})"
            qty_cell = detail_sheet.cell(row=total_row_idx, column=9, value=sum_formula)
            qty_cell.font = STYLE_FONT_BOLD
            qty_cell.alignment = ALIGN_RIGHT
            qty_cell.number_format = "#,##0"

            # 設定合計行底線樣式
            for c_idx in range(1, len(DETAIL_EXPORT_HEADERS) + 1):
                cell = detail_sheet.cell(row=total_row_idx, column=c_idx)
                cell.border = STYLE_BORDER_TOTAL
                cell.fill = STYLE_FILL_TOTAL

        _auto_fit_columns(detail_sheet)

        # 3. 統計數據分析頁
        stats_sheet = workbook.create_sheet("統計")
        stats_sheet.views.sheetView[0].showGridLines = True
        
        current_row = 1
        rows_by_name = {
            "product_name": product_rows,
            "supplier_name": supplier_rows,
            "outsource_supplier_name": outsource_rows,
        }

        for section_title, name_key in STATS_SECTION_DEFINITIONS:
            rows = rows_by_name[name_key]
            
            # 分區大標題
            stats_sheet.cell(row=current_row, column=1, value=f"📋 {section_title}").font = STYLE_FONT_BOLD
            stats_sheet.row_dimensions[current_row].height = 22
            current_row += 1
            
            # 分區表頭
            headers = build_stats_headers(name_key)
            stats_sheet.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = stats_sheet.cell(row=current_row, column=col_idx)
                cell.font = STYLE_HEADER_FONT
                cell.fill = STYLE_FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = STYLE_BORDER_THIN
            stats_sheet.row_dimensions[current_row].height = 22
            
            current_row += 1
            
            # 數據寫入
            section_start_row = current_row
            for r_idx, row in enumerate(rows, start=section_start_row):
                data = [row.get(name_key, "")] + [row.get(field_name, 0) for field_name, _ in STATS_DIMENSION_COLUMNS]
                stats_sheet.append(data)
                
                is_even = (r_idx % 2 == 0)
                for col_idx in range(1, len(headers) + 1):
                    cell = stats_sheet.cell(row=r_idx, column=col_idx)
                    cell.font = STYLE_FONT
                    cell.border = STYLE_BORDER_THIN
                    if is_even:
                        cell.fill = STYLE_FILL_ZEBRA
                        
                    # 排版格式
                    if col_idx == 1:
                        cell.alignment = ALIGN_LEFT
                    else:
                        cell.alignment = ALIGN_RIGHT
                        field_name = STATS_DIMENSION_COLUMNS[col_idx - 2][0]
                        if field_name in ("case_count", "total_qty"):
                            cell.number_format = "#,##0"
                stats_sheet.row_dimensions[r_idx].height = 20
                current_row += 1
                
            # 分區合計列
            if len(rows) > 0:
                stats_sheet.cell(row=current_row, column=1, value="合計").font = STYLE_FONT_BOLD
                stats_sheet.cell(row=current_row, column=1).alignment = ALIGN_CENTER
                
                for col_idx in range(2, len(headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    sum_formula = f"=SUM({col_letter}{section_start_row}:{col_letter}{current_row - 1})"
                    cell = stats_sheet.cell(row=current_row, column=col_idx, value=sum_formula)
                    cell.font = STYLE_FONT_BOLD
                    cell.alignment = ALIGN_RIGHT
                    
                    field_name = STATS_DIMENSION_COLUMNS[col_idx - 2][0]
                    if field_name in ("case_count", "total_qty"):
                        cell.number_format = "#,##0"
                
                for col_idx in range(1, len(headers) + 1):
                    cell = stats_sheet.cell(row=current_row, column=col_idx)
                    cell.border = STYLE_BORDER_TOTAL
                    cell.fill = STYLE_FILL_TOTAL
                    
                stats_sheet.row_dimensions[current_row].height = 20
                current_row += 1
                
            current_row += 2  # 空兩行，分隔下一個區塊

        _auto_fit_columns(stats_sheet)

        # 輸出並存檔
        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return True, f"已匯出至：{output_path}"
    except Exception as exc:
        logger.exception("匯出不合格品 Excel 報表失敗")
        return False, f"匯出不合格品 Excel 報表失敗：{exc}"


def export_to_excel(
    defects,
    product_stats,
    supplier_stats,
    outsource_stats,
    file_path: str | None = None,
) -> str:
    """向下相容用之傳統匯出方法。"""
    defect_rows = _normalize_rows(defects)
    product_rows = _normalize_rows(product_stats)
    supplier_rows = _normalize_rows(supplier_stats)
    outsource_rows = _normalize_rows(outsource_stats)

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "不良品明細"
    detail_sheet.append(DETAIL_EXPORT_HEADERS)

    for row in defect_rows:
        detail_sheet.append([row.get(field_name, "") for field_name, _ in DETAIL_EXPORT_COLUMNS])
    _auto_fit_columns(detail_sheet)

    stats_sheet = workbook.create_sheet("統計")
    current_row = 1
    rows_by_name = {
        "product_name": product_rows,
        "supplier_name": supplier_rows,
        "outsource_supplier_name": outsource_rows,
    }

    for section_title, name_key in STATS_SECTION_DEFINITIONS:
        rows = rows_by_name[name_key]
        stats_sheet.cell(row=current_row, column=1, value=section_title)
        current_row += 1
        stats_sheet.append(build_stats_headers(name_key))
        current_row += 1
        for row in rows:
            stats_sheet.append(
                [row.get(name_key, "")]
                + [row.get(field_name, 0) for field_name, _ in STATS_DIMENSION_COLUMNS]
            )
            current_row += 1
        current_row += 1

    _auto_fit_columns(stats_sheet)

    if file_path:
        output_path = Path(file_path)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = BASE_DIR / f"defect_report_{timestamp}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return str(output_path)
