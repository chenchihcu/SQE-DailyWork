from __future__ import annotations

import sqlite3
import tempfile
import unittest
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ncr.db import crud, database
from ncr.models.defect import DETAIL_EXPORT_HEADERS, build_stats_headers
from ncr.services import defect_service, export_service, product_import_service, stats_service


class DatabaseTestCase(unittest.TestCase):
    def create_memory_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        database.apply_schema(conn, with_version=True)
        return conn

    def sample_payload(self, **overrides) -> dict[str, object]:
        payload = {
            "event_date": "2026-04-14",
            "return_slip_type": "廠內退料",
            "work_order_no": "5102-260414001",
            "item_no": "ITEM-1001",
            "product_name": "Motor",
            "qty": 5,
            "category": "成品",
            "supplier_name": "A Supplier",
            "outsource_supplier_name": "Outsource A",
            "defect_desc": "Scratch on housing",
            "status": "處理中",
            "disposition": "重工",
        }
        payload.update(overrides)
        return payload


class ProductImportServiceTests(DatabaseTestCase):
    def write_product_workbook(
        self,
        path: Path,
        *,
        headers: list[str] | None = None,
        rows: list[tuple[object, ...]] | None = None,
    ) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        if headers is not None:
            worksheet.append(headers)
        for row in rows or []:
            worksheet.append(row)
        workbook.save(path)

    def count_products(self, conn: sqlite3.Connection) -> int:
        row = conn.execute("SELECT COUNT(*) FROM product_records").fetchone()
        return int(row[0])

    def test_valid_preview_does_not_write_until_confirmed_apply(self) -> None:
        conn = self.create_memory_connection()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                workbook_path = Path(temp_dir) / "products.xlsx"
                self.write_product_workbook(
                    workbook_path,
                    headers=["料號", "產品名稱"],
                    rows=[
                        ("ITEM-IMPORT-001", "Product A"),
                        ("ITEM-IMPORT-002", "Product B"),
                    ],
                )

                preview = product_import_service.preview_product_import(
                    conn, workbook_path
                )

                self.assertTrue(preview.can_import)
                self.assertEqual(preview.importable_count, 2)
                self.assertEqual(self.count_products(conn), 0)

                result = product_import_service.apply_product_import(
                    conn, preview
                )

                self.assertEqual(result.added_count, 2)
                self.assertEqual(result.product_updated_count, 0)
                self.assertEqual(result.defect_updated_count, 0)
                self.assertEqual(self.count_products(conn), 2)
        finally:
            conn.close()

    def test_invalid_import_inputs_are_blocked_and_do_not_write(self) -> None:
        scenarios = [
            ("missing_item_header", ["品號", "產品名稱"], [("ITEM-001", "Product A")]),
            ("missing_name_header", ["料號", "品名"], [("ITEM-001", "Product A")]),
            ("empty_workbook", None, []),
            ("blank_item_no", ["料號", "產品名稱"], [("", "Product A")]),
            ("blank_product_name", ["料號", "產品名稱"], [("ITEM-001", "")]),
            (
                "duplicate_item_no",
                ["料號", "產品名稱"],
                [("ITEM-001", "Product A"), ("ITEM-001", "Product B")],
            ),
        ]

        for name, headers, rows in scenarios:
            with self.subTest(name=name):
                conn = self.create_memory_connection()
                try:
                    initial_count = self.count_products(conn)

                    with tempfile.TemporaryDirectory() as temp_dir:
                        workbook_path = Path(temp_dir) / "products.xlsx"
                        self.write_product_workbook(
                            workbook_path,
                            headers=headers,
                            rows=rows,
                        )

                        preview = product_import_service.preview_product_import(
                            conn, workbook_path
                        )

                        self.assertFalse(preview.can_import)
                        self.assertGreater(preview.error_count, 0)
                        with self.assertRaises(ValueError):
                            product_import_service.apply_product_import(
                                conn, preview
                            )
                        self.assertEqual(self.count_products(conn), initial_count)
                finally:
                    conn.close()

    def test_existing_same_product_is_skipped_without_error(self) -> None:
        conn = self.create_memory_connection()
        try:
            conn.execute(
                """
                INSERT INTO product_records (item_no, product_name, created_at)
                VALUES (?, ?, ?)
                """,
                ("ITEM-001", "Product A", "2026-04-14T09:00:00"),
            )
            conn.commit()

            with tempfile.TemporaryDirectory() as temp_dir:
                workbook_path = Path(temp_dir) / "products.xlsx"
                self.write_product_workbook(
                    workbook_path,
                    headers=["item_no", "product_name"],
                    rows=[
                        ("ITEM-001", "Product A"),
                        ("ITEM-002", "Product B"),
                    ],
                )

                preview = product_import_service.preview_product_import(
                    conn, workbook_path
                )

                self.assertTrue(preview.can_import)
                self.assertEqual(preview.skipped_count, 1)
                self.assertEqual(preview.importable_count, 1)

                result = product_import_service.apply_product_import(
                    conn, preview
                )

                self.assertEqual(result.added_count, 1)
                self.assertEqual(result.product_updated_count, 0)
                self.assertEqual(result.defect_updated_count, 0)
                self.assertEqual(self.count_products(conn), 2)
        finally:
            conn.close()

    def test_existing_different_product_name_updates_products_and_defects(self) -> None:
        conn = self.create_memory_connection()
        try:
            conn.execute(
                """
                INSERT INTO product_records (item_no, product_name, created_at)
                VALUES (?, ?, ?)
                """,
                ("ITEM-001", "Old Product", "2026-04-14T09:00:00"),
            )
            for defect_no, work_order_no, product_name in [
                ("NCR-PRODUCT-001", "5102-260414001", "Old Product"),
                ("NCR-PRODUCT-002", "5102-260414002", "Another Old Product"),
            ]:
                conn.execute(
                    """
                    INSERT INTO defect_records (
                        defect_no,
                        event_date,
                        work_order_no,
                        item_no,
                        product_name,
                        qty,
                        defect_desc,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        defect_no,
                        "2026-04-14",
                        work_order_no,
                        "ITEM-001",
                        product_name,
                        1,
                        "Product name cleanup",
                        "2026-04-14T09:00:00",
                    ),
                )
            conn.commit()

            with tempfile.TemporaryDirectory() as temp_dir:
                workbook_path = Path(temp_dir) / "products.xlsx"
                self.write_product_workbook(
                    workbook_path,
                    headers=["料號", "產品名稱"],
                    rows=[
                        ("ITEM-001", "New Product"),
                        ("ITEM-002", "Product B"),
                    ],
                )

                preview = product_import_service.preview_product_import(
                    conn, workbook_path
                )

                self.assertTrue(preview.can_import)
                self.assertEqual(preview.error_count, 0)
                self.assertEqual(preview.importable_count, 1)
                self.assertEqual(preview.product_update_count, 1)
                self.assertEqual(preview.defect_update_count, 2)

                result = product_import_service.apply_product_import(conn, preview)

                updated = crud.get_product_by_item_no(conn, "ITEM-001")
                self.assertIsNotNone(updated)
                assert updated is not None
                self.assertEqual(updated["product_name"], "New Product")
                self.assertIsNotNone(crud.get_product_by_item_no(conn, "ITEM-002"))
                self.assertEqual(result.added_count, 1)
                self.assertEqual(result.product_updated_count, 1)
                self.assertEqual(result.defect_updated_count, 2)
                defect_names = {
                    row["product_name"]
                    for row in conn.execute(
                        """
                        SELECT product_name
                        FROM defect_records
                        WHERE item_no = ?
                        """,
                        ("ITEM-001",),
                    ).fetchall()
                }
                self.assertEqual(defect_names, {"New Product"})
        finally:
            conn.close()

    def test_apply_rolls_back_when_database_insert_fails(self) -> None:
        conn = self.create_memory_connection()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                workbook_path = Path(temp_dir) / "products.xlsx"
                self.write_product_workbook(
                    workbook_path,
                    headers=["料號", "產品名稱"],
                    rows=[
                        ("ITEM-ROLLBACK-001", "Product A"),
                        ("ITEM-ROLLBACK-002", "Product B"),
                    ],
                )

                preview = product_import_service.preview_product_import(
                    conn, workbook_path
                )
                self.assertTrue(preview.can_import)

                conn.execute(
                    """
                    INSERT INTO product_records (item_no, product_name, created_at)
                    VALUES (?, ?, ?)
                    """,
                    ("ITEM-ROLLBACK-002", "Product B", "2026-04-14T09:00:00"),
                )
                conn.commit()

                with self.assertRaises(sqlite3.IntegrityError):
                    product_import_service.apply_product_import(conn, preview)

                self.assertIsNone(
                    crud.get_product_by_item_no(conn, "ITEM-ROLLBACK-001")
                )
                self.assertIsNotNone(
                    crud.get_product_by_item_no(conn, "ITEM-ROLLBACK-002")
                )
        finally:
            conn.close()

    def test_apply_rolls_back_product_and_defect_updates_on_failure(self) -> None:
        conn = self.create_memory_connection()
        original_update_defects = crud.update_defect_product_names_by_item_no
        try:
            conn.execute(
                """
                INSERT INTO product_records (item_no, product_name, created_at)
                VALUES (?, ?, ?)
                """,
                ("ITEM-ROLLBACK-UPDATE", "Old Product", "2026-04-14T09:00:00"),
            )
            conn.execute(
                """
                INSERT INTO defect_records (
                    defect_no,
                    event_date,
                    work_order_no,
                    item_no,
                    product_name,
                    qty,
                    defect_desc,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "NCR-ROLLBACK-UPDATE",
                    "2026-04-14",
                    "5102-260414001",
                    "ITEM-ROLLBACK-UPDATE",
                    "Old Product",
                    1,
                    "Rollback cleanup",
                    "2026-04-14T09:00:00",
                ),
            )
            conn.commit()

            with tempfile.TemporaryDirectory() as temp_dir:
                workbook_path = Path(temp_dir) / "products.xlsx"
                self.write_product_workbook(
                    workbook_path,
                    headers=["料號", "產品名稱"],
                    rows=[("ITEM-ROLLBACK-UPDATE", "New Product")],
                )
                preview = product_import_service.preview_product_import(
                    conn,
                    workbook_path,
                )

                def fail_defect_update(
                    _conn: sqlite3.Connection,
                    _rows: list[dict[str, object]],
                ) -> int:
                    raise sqlite3.DatabaseError("forced defect update failure")

                crud.update_defect_product_names_by_item_no = fail_defect_update
                with self.assertRaisesRegex(
                    sqlite3.DatabaseError,
                    "forced defect update failure",
                ):
                    product_import_service.apply_product_import(conn, preview)

                product = crud.get_product_by_item_no(conn, "ITEM-ROLLBACK-UPDATE")
                self.assertIsNotNone(product)
                assert product is not None
                self.assertEqual(product["product_name"], "Old Product")
                defect = conn.execute(
                    """
                    SELECT product_name
                    FROM defect_records
                    WHERE item_no = ?
                    """,
                    ("ITEM-ROLLBACK-UPDATE",),
                ).fetchone()
                self.assertIsNotNone(defect)
                assert defect is not None
                self.assertEqual(defect["product_name"], "Old Product")
        finally:
            crud.update_defect_product_names_by_item_no = original_update_defects
            conn.close()

    def test_file_database_import_creates_backup_before_write(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "data" / "defect.db"
            db_path.parent.mkdir(parents=True, exist_ok=True)
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                database.apply_schema(conn, with_version=True)
                workbook_path = Path(temp_dir) / "products.xlsx"
                self.write_product_workbook(
                    workbook_path,
                    headers=["料號", "產品名稱"],
                    rows=[("ITEM-BACKUP-001", "Backup Product")],
                )
                preview = product_import_service.preview_product_import(
                    conn,
                    workbook_path,
                )

                result = product_import_service.apply_product_import(conn, preview)

                self.assertIsNotNone(result.backup_path)
                assert result.backup_path is not None
                self.assertTrue(result.backup_path.exists())
                self.assertEqual(result.backup_path.parent.name, "backups")

                backup_conn = sqlite3.connect(result.backup_path)
                try:
                    backup_count = backup_conn.execute(
                        "SELECT COUNT(*) FROM product_records"
                    ).fetchone()[0]
                finally:
                    backup_conn.close()
                self.assertEqual(backup_count, 0)
                self.assertEqual(self.count_products(conn), 1)
            finally:
                conn.close()


class InitializeDatabaseTests(DatabaseTestCase):
    def test_initialize_database_uses_shared_sqe_dailywork_schema(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            original_path = database.DB_PATH
            database.DB_PATH = Path(temp_dir) / "sqe_v2.db"
            conn: sqlite3.Connection | None = None
            try:
                conn = database.initialize_database()
                self.assertTrue(database.DB_PATH.exists())
                objects = {
                    (row["type"], row["name"])
                    for row in conn.execute(
                        """
                        SELECT type, name
                        FROM sqlite_master
                        WHERE type IN ('table', 'view')
                        """
                    ).fetchall()
                }

                for table_name in [
                    "suppliers",
                    "products",
                    "visits",
                    "visit_product_sections",
                    "visit_defect_notes",
                    "anomalies",
                    "defect_records",
                    "import_batches",
                    "import_batch_rows",
                    "ui_settings",
                ]:
                    self.assertIn(("table", table_name), objects)
                self.assertIn(("view", "supplier_records"), objects)
                self.assertIn(("view", "product_records"), objects)

                main_db = conn.execute("PRAGMA database_list").fetchone()
                self.assertIsNotNone(main_db)
                assert main_db is not None
                self.assertEqual(Path(str(main_db["file"])).name, "sqe_v2.db")
                self.assertFalse((Path(temp_dir) / "data" / "defect.db").exists())
            finally:
                if conn is not None:
                    conn.close()
                database.DB_PATH = original_path

    def test_apply_schema_still_supports_memory_core_tests(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        try:
            database.apply_schema(conn, with_version=True)
            table_names = {
                row["name"]
                for row in conn.execute(
                    """
                    SELECT name
                    FROM sqlite_master
                    WHERE type = 'table'
                    """
                ).fetchall()
            }

            self.assertIn("defect_records", table_names)
            self.assertIn("supplier_records", table_names)
            self.assertIn("product_records", table_names)
            self.assertIn("ui_settings", table_names)
            self.assertEqual(
                conn.execute("PRAGMA user_version").fetchone()[0],
                database.SCHEMA_VERSION,
            )
        finally:
            conn.close()


class DefectServiceTests(DatabaseTestCase):
    def test_generate_defect_no_increments_globally(self) -> None:
        conn = self.create_memory_connection()
        first_no = defect_service.create_defect(conn, self.sample_payload())
        second_no = defect_service.create_defect(conn, self.sample_payload(qty=2, item_no="ITEM-1002"))

        self.assertEqual(first_no, "NCR-10001")
        self.assertEqual(second_no, "NCR-10002")
        conn.close()

    def test_get_defects_supports_dynamic_filters(self) -> None:
        conn = self.create_memory_connection()
        first = defect_service.validate_defect_data(self.sample_payload())
        first["defect_no"] = "NCR-20260414-001"
        first["created_at"] = "2026-04-14T08:00:00"
        second = defect_service.validate_defect_data(
            self.sample_payload(
                event_date="2026-04-15",
                work_order_no="5102-260415001",
                item_no="ITEM-2001",
                supplier_name="B Supplier",
            )
        )
        second["defect_no"] = "NCR-20260415-001"
        second["created_at"] = "2026-04-15T08:00:00"

        crud.insert_defect(conn, first)
        crud.insert_defect(conn, second)

        results = crud.get_defects(conn, {"supplier_name": "B Supplier"})
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["return_slip_type"], "廠內退料")
        self.assertEqual(results[0]["work_order_no"], "5102-260415001")
        conn.close()

    def test_validate_defect_data_rejects_non_iso_event_date(self) -> None:
        with self.assertRaisesRegex(ValueError, "YYYY-MM-DD"):
            defect_service.validate_defect_data(
                self.sample_payload(event_date="2026/04/14")
            )

    def test_validate_defect_data_rejects_future_event_date(self) -> None:
        future_date = (date.today() + timedelta(days=1)).isoformat()
        with self.assertRaisesRegex(ValueError, "不可晚於今天"):
            defect_service.validate_defect_data(
                self.sample_payload(event_date=future_date)
            )

    def test_validate_defect_data_defaults_status_when_required(self) -> None:
        normalized = defect_service.validate_defect_data(
            self.sample_payload(status=""),
            require_status_default=True,
        )
        self.assertEqual(normalized["status"], "處理中")

    def test_validate_defect_data_requires_return_slip_type(self) -> None:
        with self.assertRaisesRegex(ValueError, "退料單別為必填"):
            defect_service.validate_defect_data(self.sample_payload(return_slip_type=""))

    def test_validate_defect_data_rejects_invalid_return_slip_type(self) -> None:
        with self.assertRaisesRegex(ValueError, "退料單別選項不正確"):
            defect_service.validate_defect_data(
                self.sample_payload(return_slip_type="其他")
            )

    def test_validate_defect_data_accepts_outsource_return_slip_type(self) -> None:
        normalized = defect_service.validate_defect_data(
            self.sample_payload(return_slip_type="託外退料")
        )
        self.assertEqual(normalized["return_slip_type"], "託外退料")

    def test_validate_defect_data_rejects_empty_status_without_default(self) -> None:
        with self.assertRaisesRegex(ValueError, "狀態選項不正確"):
            defect_service.validate_defect_data(
                self.sample_payload(status=""),
                require_status_default=False,
            )

    def test_validate_defect_data_rejects_removed_dispositions(self) -> None:
        for removed_option in ("退貨", "換料", "特採"):
            with self.subTest(disposition=removed_option):
                with self.assertRaisesRegex(ValueError, "處置方式選項不正確"):
                    defect_service.validate_defect_data(
                        self.sample_payload(disposition=removed_option)
                    )

    def test_validate_defect_data_allows_optional_freeform_work_orders(self) -> None:
        # 委外製令 / 廠內製令 are optional free-text fields: blank or any value saves.
        normalized = defect_service.validate_defect_data(
            self.sample_payload(work_order_no="", internal_work_order_no="")
        )
        self.assertEqual(normalized["work_order_no"], "")
        self.assertEqual(normalized["internal_work_order_no"], "")

        normalized = defect_service.validate_defect_data(
            self.sample_payload(
                work_order_no="ABC-123",
                internal_work_order_no="任意文字",
            )
        )
        self.assertEqual(normalized["work_order_no"], "ABC-123")
        self.assertEqual(normalized["internal_work_order_no"], "任意文字")

        # Previously-valid formatted values still pass; values are trimmed.
        normalized = defect_service.validate_defect_data(
            self.sample_payload(work_order_no="  5102-260414001  ")
        )
        self.assertEqual(normalized["work_order_no"], "5102-260414001")

    def test_create_defect_allows_blank_work_order(self) -> None:
        conn = self.create_memory_connection()
        try:
            defect_no = defect_service.create_defect(
                conn, self.sample_payload(work_order_no="", internal_work_order_no="")
            )
            row = conn.execute(
                "SELECT work_order_no, internal_work_order_no "
                "FROM defect_records WHERE defect_no = ?",
                (defect_no,),
            ).fetchone()
            assert row is not None
            self.assertEqual(row["work_order_no"], "")
            self.assertEqual(row["internal_work_order_no"], "")
        finally:
            conn.close()

    def test_create_defect_blocks_duplicate_business_key(self) -> None:
        conn = self.create_memory_connection()
        defect_service.create_defect(conn, self.sample_payload())

        with self.assertRaisesRegex(ValueError, "已有相同發生日期"):
            defect_service.create_defect(conn, self.sample_payload(qty=9, status="處理中"))
        conn.close()

    def test_update_defect_blocks_duplicate_business_key(self) -> None:
        conn = self.create_memory_connection()
        defect_service.create_defect(conn, self.sample_payload())
        defect_service.create_defect(
            conn,
            self.sample_payload(
                work_order_no="5102-260414002",
                item_no="ITEM-1002",
                defect_desc="Scratch B",
            ),
        )

        second_row = conn.execute(
            "SELECT id FROM defect_records WHERE work_order_no = ?",
            ("5102-260414002",),
        ).fetchone()
        assert second_row is not None

        with self.assertRaisesRegex(ValueError, "已有相同發生日期"):
            defect_service.update_defect(
                conn,
                int(second_row["id"]),
                self.sample_payload(),
            )
        conn.close()

    def test_stats_service_returns_expected_totals(self) -> None:
        conn = self.create_memory_connection()
        defect_service.create_defect(
            conn,
            self.sample_payload(
                product_name="Motor",
                qty=5,
                category="成品",
                status="處理中",
                disposition="重工",
                item_no="ITEM-1001-A",
                work_order_no="5102-260414003",
                defect_desc="desc-1",
            ),
        )
        defect_service.create_defect(
            conn,
            self.sample_payload(
                work_order_no="5102-260414002",
                item_no="ITEM-1002",
                product_name="Motor",
                qty=3,
                supplier_name="A Supplier",
                outsource_supplier_name="Outsource B",
                category="原物料",
                status="處理中",
                disposition="重工",
                defect_desc="desc-2",
            ),
        )
        defect_service.create_defect(
            conn,
            self.sample_payload(
                work_order_no="5102-260414004",
                item_no="ITEM-1003",
                product_name="Motor",
                qty=2,
                supplier_name="A Supplier",
                outsource_supplier_name="Outsource B",
                category="原物料",
                status="處理中",
                disposition="報廢",
                defect_desc="desc-3",
            ),
        )

        product_stats = stats_service.get_product_stats(conn)
        supplier_stats = stats_service.get_supplier_stats(conn)
        outsource_stats = stats_service.get_outsource_stats(conn)
        status_totals = stats_service.get_status_totals(conn)
        outsource_processing_stats = stats_service.get_outsource_processing_stats(conn)
        supplier_processing_stats = stats_service.get_supplier_processing_stats(conn)
        outsource_scrap_stats = stats_service.get_outsource_scrap_stats(conn)
        outsource_processing_preview_rows = (
            stats_service.get_outsource_processing_preview_rows(conn)
        )
        supplier_processing_preview_rows = (
            stats_service.get_supplier_processing_preview_rows(conn)
        )
        outsource_scrap_preview_rows = stats_service.get_outsource_scrap_preview_rows(conn)
        warehouse_summary = stats_service.get_warehouse_nonconforming_summary(conn)

        self.assertEqual(product_stats[0]["product_name"], "Motor")
        self.assertEqual(product_stats[0]["disposition"], "重工")
        self.assertEqual(product_stats[0]["category"], "成品")
        self.assertEqual(product_stats[0]["event_month"], "2026-04")
        self.assertEqual(product_stats[0]["status"], "處理中")
        self.assertEqual(product_stats[0]["case_count"], 1)
        self.assertEqual(product_stats[0]["total_qty"], 5)
        self.assertEqual(supplier_stats[0]["supplier_name"], "A Supplier")
        self.assertEqual(supplier_stats[0]["event_month"], "2026-04")
        self.assertEqual(outsource_stats[0]["case_count"], 1)
        status_total_by_name = {row["status"]: row["total_qty"] for row in status_totals}
        self.assertEqual(status_total_by_name["處理中"], 10)  # 5 + 3 + 2
        self.assertEqual(len(outsource_processing_stats), 2)
        self.assertEqual(outsource_processing_stats[0]["outsource_supplier_name"], "Outsource A")
        self.assertEqual(outsource_processing_stats[0]["total_qty"], 5)
        self.assertEqual(len(supplier_processing_stats), 1)
        self.assertEqual(supplier_processing_stats[0]["supplier_name"], "A Supplier")
        self.assertEqual(supplier_processing_stats[0]["total_qty"], 10)
        self.assertEqual(len(outsource_scrap_stats), 1)
        self.assertEqual(outsource_scrap_stats[0]["outsource_supplier_name"], "Outsource B")
        self.assertEqual(outsource_scrap_stats[0]["item_no"], "ITEM-1003")
        self.assertEqual(outsource_scrap_stats[0]["total_qty"], 2)
        self.assertEqual(len(outsource_processing_preview_rows), 3)
        self.assertEqual(
            outsource_processing_preview_rows[0]["outsource_supplier_name"], "Outsource B"
        )
        self.assertEqual(outsource_processing_preview_rows[0]["status"], "處理中")
        self.assertEqual(len(supplier_processing_preview_rows), 3)
        self.assertEqual(supplier_processing_preview_rows[0]["supplier_name"], "A Supplier")
        self.assertEqual(supplier_processing_preview_rows[0]["status"], "處理中")
        self.assertEqual(len(outsource_scrap_preview_rows), 1)
        self.assertEqual(outsource_scrap_preview_rows[0]["outsource_supplier_name"], "Outsource B")
        self.assertEqual(outsource_scrap_preview_rows[0]["item_no"], "ITEM-1003")
        self.assertEqual(outsource_scrap_preview_rows[0]["disposition"], "報廢")
        self.assertEqual(3, warehouse_summary["total_count"])
        self.assertEqual(10, warehouse_summary["total_qty"])
        self.assertEqual(3, warehouse_summary["open_count"])
        self.assertEqual(10, warehouse_summary["open_qty"])
        self.assertEqual(0, warehouse_summary["closed_count"])
        self.assertEqual(8, warehouse_summary["rework_qty"])
        self.assertEqual(2, warehouse_summary["scrap_qty"])
        conn.close()

    def test_stats_preview_handles_single_quote_in_filter_values(self) -> None:
        conn = self.create_memory_connection()
        conn.execute(
            """
            INSERT INTO defect_records (
                defect_no,
                event_date,
                work_order_no,
                item_no,
                product_name,
                qty,
                category,
                supplier_name,
                outsource_supplier_name,
                defect_desc,
                status,
                disposition,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "NCR-20260414-001",
                "2026-04-14",
                "WO-QUOTE-001",
                "ITEM-QUOTE-001",
                "Motor",
                1,
                "成品",
                "",
                "Outsource A",
                "desc",
                "O'Reilly",
                "",
                "2026-04-14T08:00:00",
            ),
        )
        conn.commit()

        rows = stats_service._get_preview_rows_for_name(
            conn,
            name_field="outsource_supplier_name",
            status="O'Reilly",
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["status"], "O'Reilly")
        conn.close()

    def test_stats_service_rejects_unsupported_dynamic_name_field(self) -> None:
        conn = self.create_memory_connection()
        with self.assertRaisesRegex(ValueError, "Unsupported stats name field"):
            stats_service._get_preview_rows_for_name(
                conn,
                name_field="status",
                status="處理中",
            )
        conn.close()


class DatabaseConstraintTests(DatabaseTestCase):
    def test_future_date_trigger_blocks_insert_and_update(self) -> None:
        conn = self.create_memory_connection()
        tomorrow = (date.today() + timedelta(days=1)).isoformat()

        with self.assertRaisesRegex(sqlite3.IntegrityError, "event_date cannot be in future"):
            conn.execute(
                """
                INSERT INTO defect_records (
                    defect_no,
                    event_date,
                    work_order_no,
                    item_no,
                    product_name,
                    qty,
                    category,
                    supplier_name,
                    outsource_supplier_name,
                    defect_desc,
                    status,
                    disposition,
                    created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "NCR-99990101-001",
                    tomorrow,
                    "WO-X",
                    "ITEM-X",
                    "Product",
                    1,
                    "成品",
                    "",
                    "",
                    "desc",
                    "處理中",
                    "",
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )

        defect_no = defect_service.create_defect(conn, self.sample_payload())
        row = conn.execute(
            "SELECT id FROM defect_records WHERE defect_no = ?",
            (defect_no,),
        ).fetchone()
        assert row is not None

        with self.assertRaisesRegex(sqlite3.IntegrityError, "event_date cannot be in future"):
            conn.execute(
                "UPDATE defect_records SET event_date = ? WHERE id = ?",
                (tomorrow, int(row["id"])),
            )
        conn.close()


class CrudSafeguardTests(DatabaseTestCase):
    def test_update_defect_raises_for_missing_id(self) -> None:
        conn = self.create_memory_connection()
        normalized = defect_service.validate_defect_data(self.sample_payload())
        with self.assertRaisesRegex(sqlite3.DatabaseError, "找不到要更新的資料 ID"):
            crud.update_defect(conn, 9999, normalized)
        conn.close()

    def test_delete_defect_raises_for_missing_id(self) -> None:
        conn = self.create_memory_connection()
        with self.assertRaisesRegex(sqlite3.DatabaseError, "找不到要刪除的資料 ID"):
            crud.delete_defect(conn, 9999)
        conn.close()


class ExportServiceTests(DatabaseTestCase):
    def test_export_to_excel_creates_expected_workbook(self) -> None:
        defects = [
            {
                "defect_no": "NCR-20260414-001",
                "event_date": "2026-04-14",
                "return_slip_type": "廠內退料",
                "work_order_no": "5102-260414001",
                "item_no": "ITEM-1001",
                "product_name": "Motor",
                "qty": 5,
                "category": "成品",
                "supplier_name": "A Supplier",
                "outsource_supplier_name": "Outsource A",
                "defect_desc": "Scratch on housing",
                "status": "處理中",
                "disposition": "重工",
            }
        ]
        product_stats = [
            {
                "product_name": "Motor",
                "disposition": "重工",
                "category": "成品",
                "event_month": "2026-04",
                "status": "處理中",
                "case_count": 1,
                "total_qty": 5,
            }
        ]
        supplier_stats = [
            {
                "supplier_name": "A Supplier",
                "disposition": "重工",
                "category": "成品",
                "event_month": "2026-04",
                "status": "處理中",
                "case_count": 1,
                "total_qty": 5,
            }
        ]
        outsource_stats = [
            {
                "outsource_supplier_name": "Outsource A",
                "disposition": "重工",
                "category": "成品",
                "event_month": "2026-04",
                "status": "處理中",
                "case_count": 1,
                "total_qty": 5,
            }
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "report.xlsx"
            result = export_service.export_to_excel(
                defects,
                product_stats,
                supplier_stats,
                outsource_stats,
                file_path=str(file_path),
            )
            self.assertEqual(result, str(file_path))
            self.assertTrue(file_path.exists())

            workbook = load_workbook(file_path)
            self.assertEqual(workbook.sheetnames, ["不良品明細", "統計"])
            detail_headers = [
                workbook["不良品明細"].cell(row=1, column=index).value
                for index in range(1, len(DETAIL_EXPORT_HEADERS) + 1)
            ]
            stats_headers = [
                workbook["統計"].cell(row=2, column=index).value
                for index in range(1, len(build_stats_headers("product_name")) + 1)
            ]
            self.assertEqual(detail_headers, DETAIL_EXPORT_HEADERS)
            self.assertEqual(stats_headers, build_stats_headers("product_name"))
            self.assertEqual(workbook["不良品明細"]["A2"].value, "NCR-20260414-001")
            self.assertEqual(workbook["不良品明細"]["C2"].value, "廠內退料")
            self.assertEqual(workbook["統計"]["A1"].value, "產品統計")
            self.assertEqual(workbook["統計"]["B2"].value, "處置方式")
            self.assertEqual(workbook["統計"]["D3"].value, "2026-04")


class DefectRecordsWorkOrderMigrationTests(unittest.TestCase):
    """Existing DBs with the legacy work_order_no CHECK must be rebuilt as optional."""

    LEGACY_DEFECT_RECORDS_DDL = """
        CREATE TABLE defect_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            defect_no TEXT NOT NULL UNIQUE CHECK(TRIM(defect_no) <> ''),
            event_date TEXT NOT NULL
                CHECK(
                    event_date GLOB '[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]'
                    AND date(event_date) IS NOT NULL
                ),
            return_slip_type TEXT NOT NULL DEFAULT '',
            work_order_no TEXT NOT NULL CHECK(TRIM(work_order_no) <> ''),
            internal_work_order_no TEXT NOT NULL DEFAULT '',
            transfer_slip_no TEXT NOT NULL DEFAULT '',
            item_no TEXT NOT NULL CHECK(TRIM(item_no) <> ''),
            product_name TEXT NOT NULL DEFAULT '',
            qty INTEGER NOT NULL CHECK(qty > 0),
            category TEXT NOT NULL DEFAULT '',
            supplier_name TEXT NOT NULL DEFAULT '',
            outsource_supplier_name TEXT NOT NULL DEFAULT '',
            defect_desc TEXT NOT NULL CHECK(TRIM(defect_desc) <> ''),
            status TEXT NOT NULL DEFAULT '',
            disposition TEXT NOT NULL DEFAULT '',
            responsibility TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL CHECK(TRIM(created_at) <> '')
        )
    """

    def _legacy_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.execute(self.LEGACY_DEFECT_RECORDS_DDL)
        conn.execute(
            """
            CREATE UNIQUE INDEX uniq_defect_records_business_key
                ON defect_records(
                    event_date, work_order_no, internal_work_order_no,
                    transfer_slip_no, item_no, defect_desc
                )
            """
        )
        conn.execute(
            """
            INSERT INTO defect_records
                (defect_no, event_date, work_order_no, item_no, qty, defect_desc, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "NCR-10001",
                "2026-04-14",
                "5102-260414001",
                "ITEM-1001",
                5,
                "Scratch",
                "2026-04-14T09:00:00",
            ),
        )
        conn.commit()
        return conn

    def test_normalizer_drops_check_and_preserves_rows(self) -> None:
        from database import repository

        conn = self._legacy_connection()
        try:
            # Legacy schema rejects a blank work order.
            with self.assertRaises(sqlite3.IntegrityError):
                conn.execute(
                    """
                    INSERT INTO defect_records
                        (defect_no, event_date, work_order_no, item_no, qty, defect_desc, created_at)
                    VALUES ('NCR-10002', '2026-04-14', '', 'ITEM-1002', 1, 'Dent', '2026-04-14T10:00:00')
                    """
                )

            repository._normalize_defect_records_optional_work_order(conn)

            table_sql = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='defect_records'"
            ).fetchone()[0]
            self.assertNotIn("CHECK(TRIM(work_order_no)", table_sql)

            # Original row preserved.
            preserved = conn.execute(
                "SELECT work_order_no FROM defect_records WHERE defect_no = 'NCR-10001'"
            ).fetchone()
            self.assertEqual(preserved["work_order_no"], "5102-260414001")

            # The unique business-key index survives the rebuild.
            index_row = conn.execute(
                "SELECT 1 FROM sqlite_master "
                "WHERE type='index' AND name='uniq_defect_records_business_key'"
            ).fetchone()
            self.assertIsNotNone(index_row)

            # A blank work order now inserts successfully.
            conn.execute(
                """
                INSERT INTO defect_records
                    (defect_no, event_date, work_order_no, item_no, qty, defect_desc, created_at)
                VALUES ('NCR-10002', '2026-04-14', '', 'ITEM-1002', 1, 'Dent', '2026-04-14T10:00:00')
                """
            )
            conn.commit()
            count = conn.execute("SELECT COUNT(*) FROM defect_records").fetchone()[0]
            self.assertEqual(int(count), 2)
        finally:
            conn.close()

    def test_normalizer_is_idempotent(self) -> None:
        from database import repository

        conn = self._legacy_connection()
        try:
            repository._normalize_defect_records_optional_work_order(conn)
            sql_after_first = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='defect_records'"
            ).fetchone()[0]
            # Second run must be a safe no-op (constraint already gone).
            repository._normalize_defect_records_optional_work_order(conn)
            sql_after_second = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='defect_records'"
            ).fetchone()[0]
            self.assertEqual(sql_after_first, sql_after_second)
        finally:
            conn.close()


if __name__ == "__main__":
    unittest.main()





