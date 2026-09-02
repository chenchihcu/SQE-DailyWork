"""Phase 7 export parity and report regression tests."""

from __future__ import annotations

import os
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest import mock

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication
from openpyxl import load_workbook

from ui.list_column_contract import MANAGER_SUMMARY_COLUMNS
from database import connection as _connection
from database import repository
from services import manager_export_service
from services import supplier_report_service
from services.event import _export_service
from services.event import _anomaly_markdown
from services.event._export_service import _append_hypothesis_export_sheet
from services.event._hypothesis_tree_png import (
    HYPOTHESIS_EXCEL_PNG_LIMIT,
    format_hypothesis_tree_text,
    render_hypothesis_tree_png,
)


def _bootstrap_db() -> tuple[str, str, str]:
    tmpdir = tempfile.mkdtemp(prefix="exports-phase7-")
    db_path = os.path.join(tmpdir, "sqe_v2.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    repository.create_schema(conn)
    supplier = repository.create_supplier_record(conn, supplier_name="匯出測試供應商")
    product = repository.create_product_record(
        conn,
        product_code="EXP-001",
        product_name="匯出測試品",
        supplier_id=supplier,
    )
    anomaly = repository.create_anomaly_with_visit_link(
        conn,
        anomaly_date="2026-08-01",
        supplier_id=supplier,
        product_id=product,
        problem_desc="匯出測試異常",
        sync_visit=False,
    )
    anomaly_id = anomaly["anomaly_id"]
    repository.create_anomaly_hypothesis(
        conn,
        anomaly_id=anomaly_id,
        statement="焊點虛焊",
        status="提案",
    )
    conn.commit()
    conn.close()
    return db_path, supplier, anomaly_id


class ExportPhase7Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def setUp(self) -> None:
        self.db_path, self.supplier_id, self.anomaly_id = _bootstrap_db()

        def _factory(*args, **kwargs):
            real = sqlite3.connect(
                self.db_path, factory=_connection.ClosingConnection
            )
            real.row_factory = sqlite3.Row
            real.execute("PRAGMA foreign_keys=ON")
            real.execute("PRAGMA journal_mode=WAL")
            real.execute("PRAGMA busy_timeout=5000")
            return real

        self._patcher = mock.patch.object(
            _connection, "get_connection", side_effect=_factory
        )
        self._patcher.start()

    def tearDown(self) -> None:
        self._patcher.stop()
        if os.path.exists(self.db_path):
            try:
                os.remove(self.db_path)
            except PermissionError:
                pass

    def test_hypothesis_export_respects_excel_png_limit(self) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        anomaly_rows = [
            {
                "event_id": f"aid-{idx:03d}",
                "ref_no": f"20260801{idx:03d}",
                "supplier_name": "匯出測試供應商",
                "hypothesis_count": 1,
            }
            for idx in range(1, HYPOTHESIS_EXCEL_PNG_LIMIT + 2)
        ]
        hypotheses = [
            {
                "id": "h1",
                "level": 1,
                "status": "提案",
                "statement": "焊點虛焊",
                "parent_hypothesis_id": "",
            }
        ]
        warnings: list[str] = []
        with tempfile.TemporaryDirectory() as tmp:

            def _write_stub_png(_hypotheses, png_path) -> bool:
                path = Path(png_path)
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(
                    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
                    b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde"
                    b"\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00"
                    b"\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
                )
                return True

            with mock.patch(
                "services.event._anomaly_workbench_service.list_hypotheses",
                return_value=hypotheses,
            ), mock.patch(
                "services.event._hypothesis_tree_png.render_hypothesis_tree_png",
                side_effect=_write_stub_png,
            ) as render_png:
                _append_hypothesis_export_sheet(
                    workbook,
                    anomaly_rows,
                    output_parent=Path(tmp),
                    warnings=warnings,
                )
            sheet = workbook["原因假設"]
            self.assertEqual(HYPOTHESIS_EXCEL_PNG_LIMIT + 1, sheet.max_row - 1)
            self.assertEqual(HYPOTHESIS_EXCEL_PNG_LIMIT, len(sheet._images))
            self.assertEqual(HYPOTHESIS_EXCEL_PNG_LIMIT, render_png.call_count)
            self.assertTrue(
                any(
                    f"僅嵌入前 {HYPOTHESIS_EXCEL_PNG_LIMIT} 案" in warning
                    for warning in warnings
                )
            )

    def test_hypothesis_tree_png_and_text_fallback(self) -> None:
        hypotheses = [
            {
                "id": "h1",
                "level": 1,
                "status": "提案",
                "statement": "焊點虛焊",
                "parent_hypothesis_id": "",
            }
        ]
        text = format_hypothesis_tree_text(hypotheses)
        self.assertIn("L1", text)
        self.assertIn("焊點虛焊", text)
        with tempfile.TemporaryDirectory() as tmp:
            png_path = Path(tmp) / "tree.png"
            self.assertTrue(render_hypothesis_tree_png(hypotheses, png_path))
            self.assertGreater(png_path.stat().st_size, 0)

    def test_export_events_report_appends_overview_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "report.xlsx")
            ok, message = _export_service.export_events_report(
                path,
                "2026-08-01",
                "2026-08-31",
                temp_chart_paths=None,
            )
            self.assertTrue(ok)
            workbook = load_workbook(path)
            anomaly_sheet = workbook["異常"]
            headers = [anomaly_sheet.cell(row=1, column=col).value for col in range(1, 32)]
            self.assertIn("原因假設數", headers)
            self.assertIn("已採納假設", headers)
            self.assertIn("重複警示", headers)
            self.assertIn("原因假設", workbook.sheetnames)

    def test_export_events_report_skips_charts_when_preference_disabled(self) -> None:
        with mock.patch(
            "services.appearance_preferences_service.load_application_preferences"
        ) as load_prefs:
            prefs = mock.Mock()
            prefs.export_include_charts = False
            prefs.excel_theme_style = "classic_navy"
            prefs.export_include_disclaimer = True
            load_prefs.return_value = prefs
            with tempfile.TemporaryDirectory() as tmp:
                path = str(Path(tmp) / "report-no-charts.xlsx")
                fake_chart = str(Path(tmp) / "fake.png")
                Path(fake_chart).write_bytes(b"not-a-real-png")
                ok, _message = _export_service.export_events_report(
                    path,
                    "2026-08-01",
                    "2026-08-31",
                    temp_chart_paths={"trend": fake_chart},
                )
                self.assertTrue(ok)
                workbook = load_workbook(path)
                self.assertNotIn("原因假設", workbook.sheetnames)

    def test_markdown_snapshot_includes_overview_block(self) -> None:
        with _connection.get_connection() as conn:
            detail = repository.get_anomaly_detail(conn, self.anomaly_id)
        text = _anomaly_markdown.render_anomaly_markdown(detail)
        self.assertIn("案件概況:", text)
        self.assertIn("原因假設數:", text)
        self.assertIn("開啟中處置:", text)

    def test_manager_export_single_summary_sheet(self) -> None:
        from services import manager_view_service

        summary_rows = manager_view_service.list_manager_summary_rows()
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "manager.xlsx")
            ok, _message = manager_export_service.export_manager_view_excel(
                path,
                summary_rows,
            )
            self.assertTrue(ok)
            workbook = load_workbook(path)
            self.assertEqual({"案件總覽"}, set(workbook.sheetnames))
            summary_sheet = workbook["案件總覽"]
            overdue_col = next(
                index + 1
                for index, column in enumerate(MANAGER_SUMMARY_COLUMNS)
                if column.field == "overdue"
            )
            overdue_values = {
                summary_sheet.cell(row=row, column=overdue_col).value
                for row in range(2, summary_sheet.max_row + 1)
            }
            self.assertTrue(overdue_values.issubset({"逾期", "—"}))
            self.assertNotIn("是", overdue_values)
            self.assertNotIn("否", overdue_values)

    def test_supplier_report_includes_repeat_and_overview_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "supplier.xlsx")
            ok, _message = supplier_report_service.export_supplier_report(
                path,
                self.supplier_id,
                "2026-08-01",
                "2026-08-31",
            )
            self.assertTrue(ok)
            workbook = load_workbook(path)
            overview = workbook["供應商摘要"]
            labels = [overview.cell(row=row, column=1).value for row in range(1, 10)]
            self.assertIn("重複警示", labels)
            anomaly_sheet = workbook["異常統計"]
            headers = [anomaly_sheet.cell(row=1, column=col).value for col in range(1, 12)]
            self.assertIn("overdue", headers)
            self.assertIn("root_cause_status", headers)

    def test_weekly_report_overdue_uses_overview_flag(self) -> None:
        import scripts.generate_weekly_report as weekly_report

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        rows = weekly_report.fetch_open_anomalies(conn)
        self.assertTrue(rows)
        self.assertIn("overview_overdue", rows[0])
        overdue, new_items, others = weekly_report.categorize_anomalies(rows)
        self.assertEqual(len(overdue) + len(new_items) + len(others), len(rows))
        conn.close()


if __name__ == "__main__":
    unittest.main()
