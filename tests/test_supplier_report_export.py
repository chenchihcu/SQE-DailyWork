from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from database.repository import create_schema
from services import supplier_report_service


class SupplierReportExportTests(unittest.TestCase):
    def test_export_creates_source_separated_sheets(self) -> None:
        conn = sqlite3.connect(":memory:")
        self.addCleanup(conn.close)
        conn.row_factory = sqlite3.Row
        create_schema(conn)
        conn.execute(
            """
            INSERT INTO suppliers(
                id, supplier_name, contact_name, department, phone,
                contact_email, category, is_active
            ) VALUES ('sup-1', '報告供應商', '', '', '', '', '原物料供應商', 1)
            """
        )
        conn.execute(
            """
            INSERT INTO anomalies(
                id, anomaly_no, anomaly_date, supplier_id, problem_desc, status
            ) VALUES ('a-1', '20260801001', '2026-08-01', 'sup-1', '異常', '待處理')
            """
        )
        conn.execute(
            """
            INSERT INTO defect_records(
                defect_no, event_date, supplier_id, supplier_name,
                processing_line, defect_desc, status, item_no, qty, created_at
            ) VALUES (
                'NCR-90001', '2026-08-03', 'sup-1', '報告供應商',
                '原物料', '不良', '待處理', 'PN-001', 1, '2026-08-03'
            )
            """
        )
        conn.commit()

        with patch.object(
            supplier_report_service.supplier_360_service._connection,
            "get_connection",
            return_value=conn,
        ):
            with tempfile.TemporaryDirectory() as tmp:
                path = str(Path(tmp) / "supplier-report.xlsx")
                ok, _message = supplier_report_service.export_supplier_report(
                    path,
                    "sup-1",
                    "2026-08-01",
                    "2026-08-31",
                )
                self.assertTrue(ok)
                from openpyxl import load_workbook

                workbook = load_workbook(path)
                self.addCleanup(workbook.close)
                titles = set(workbook.sheetnames)
                self.assertIn("供應商摘要", titles)
                self.assertIn("異常統計", titles)
                self.assertNotIn("訪廠紀錄", titles)
                self.assertIn("不合格品統計", titles)
                self.assertIn("評分摘要", titles)


if __name__ == "__main__":
    unittest.main()
