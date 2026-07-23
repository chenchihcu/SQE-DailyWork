from __future__ import annotations

import os
import sqlite3
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from uuid import uuid4

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtCore import QDate
from PySide6.QtWidgets import QApplication
from openpyxl import load_workbook

from database import repository
from ui.widgets.export_range_dialog import ExportRangeDialog
from ncr.services import export_service as ncr_export_service
from services import event_service


class ExcelReportCustomRangeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    @classmethod
    def tearDownClass(cls) -> None:
        if cls.app is not None:
            cls.app.quit()

    def test_export_range_dialog_defaults(self) -> None:
        dialog = ExportRangeDialog()
        # 預設開始日期是今年 1 月 1 日
        expected_start = QDate(QDate.currentDate().year(), 1, 1).toString("yyyy-MM-dd")
        expected_end = QDate.currentDate().toString("yyyy-MM-dd")
        
        start_val, end_val = dialog.get_date_range()
        self.assertEqual(start_val, expected_start)
        self.assertEqual(end_val, expected_end)
        dialog.close()

    def test_ncr_report_export_success(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            file_path = os.path.join(tmp_dir, "ncr_report.xlsx")
            
            # 建立虛擬不合格品資料
            defects = [
                {
                    "defect_no": "NCR-10001",
                    "event_date": "2026-06-15",
                    "return_slip_type": "廠內退料",
                    "work_order_no": "WO-12345",
                    "internal_work_order_no": "IWO-54321",
                    "transfer_slip_no": "TS-999",
                    "item_no": "ITEM-ABC",
                    "product_name": "虛擬產品A",
                    "qty": 50,
                    "category": "原物料",
                    "supplier_name": "供應商A",
                    "outsource_supplier_name": "N/A",
                    "defect_desc": "表面刮傷",
                    "status": "已結案",
                    "disposition": "報廢",
                    "responsibility": "材損",
                }
            ]
            
            # 呼叫匯出函數，無圖表
            ok, msg = ncr_export_service.export_ncr_excel_report(
                file_path=file_path,
                start_date="2026-06-01",
                end_date="2026-06-30",
                defects=defects,
                temp_chart_paths=None
            )
            
            self.assertTrue(ok)
            self.assertTrue(os.path.exists(file_path))
            self.assertIn("已匯出至", msg)

    @patch("services.event._query_service.get_anomaly_closure_activity_by_range")
    @patch("services.event._query_service.get_anomaly_category_pareto_by_range")
    @patch("services.event._query_service.list_events_by_range")
    def test_events_report_export_success(
        self, mock_list_events, mock_pareto, mock_closure_activity
    ) -> None:
        # 柏拉圖表格與頁面圖表、嵌入 PNG 共用同一 SQL 來源(單一實作)
        mock_pareto.return_value = [
            {
                "rank": 1,
                "category": "物料/來料品質異常",
                "count": 1,
                "percent": 100.0,
                "cumulative_percent": 100.0,
            }
        ]
        mock_closure_activity.return_value = 7
        # Mock 範圍事件列表
        mock_list_events.return_value = [
            {
                "event_id": "EVT-101",
                "ref_no": "AN-2026-001",
                "event_date": "2026-06-10",
                "event_type": "ANOMALY",
                "supplier_name": "供應商A",
                "content": "進料不合格",
                "status": "已結案",
                "category": "物料/來料品質異常",
                "category_raw": "進料品質",
                "root_cause_category": "物料/來料品質異常",
                "product_name": "控制板",
                "product_code": "PCBA-001",
                "product_stage": "試產",
                "pending_items": "追蹤 8D",
                "responsible_person": "王小明",
                "improvement_desc": "已要求廠商重工",
                "closed_at": "2026-06-12",
                "quality_report_required": True,
            },
            {
                "event_id": "EVT-103",
                "ref_no": "AN-2026-002",
                "event_date": "2026-06-11",
                "event_type": "ANOMALY",
                "supplier_name": "供應商A",
                "content": "不要求品質異常單",
                "status": "待處理",
                "category": "其他",
                "category_raw": "其他",
                "root_cause_category": "",
                "product_name": "控制板",
                "product_code": "PCBA-001",
                "product_stage": "量產",
                "pending_items": "",
                "responsible_person": "SQE",
                "improvement_desc": "",
                "closed_at": None,
                "quality_report_required": False,
            },
            {
                "event_id": "EVT-104",
                "ref_no": "AN-2026-003",
                "event_date": "2026-06-12",
                "event_type": "ANOMALY",
                "supplier_name": "供應商A",
                "content": "歷史資料",
                "status": "待處理",
                "category": "其他",
                "category_raw": "其他",
                "root_cause_category": "",
                "product_name": "",
                "product_code": "",
                "product_stage": "量產",
                "pending_items": "",
                "responsible_person": "",
                "improvement_desc": "",
                "closed_at": None,
                "quality_report_required": None,
            },
            {
                "event_id": "EVT-102",
                "ref_no": "",
                "event_date": "2026-06-15",
                "event_type": "VISIT",
                "supplier_name": "供應商B",
                "content": "例行訪廠交流",
                "status": "已完成",
                "category": "",
                "root_cause_category": "",
                "improvement_desc": "",
                "closed_at": None
            }
        ]

        with TemporaryDirectory() as tmp_dir:
            file_path = os.path.join(tmp_dir, "events_report.xlsx")
            
            ok, msg = event_service.export_events_report(
                file_path=file_path,
                start_date="2026-06-01",
                end_date="2026-06-30",
                temp_chart_paths={"category_pareto": os.path.join(tmp_dir, "missing_pareto.png")}
            )
            
            self.assertTrue(ok)
            self.assertTrue(os.path.exists(file_path))
            self.assertIn("已匯出至", msg)

            workbook = load_workbook(file_path)
            self.assertIn("異常類別柏拉圖", workbook.sheetnames)
            report_sheet = workbook["統計報告"]
            self.assertIn("期間新增異常", report_sheet["A5"].value)
            self.assertIn("其中", report_sheet["E5"].value)
            self.assertIn("目前結案率", report_sheet["F5"].value)
            self.assertIn("依結案日期", report_sheet["G5"].value)
            self.assertIn("7", report_sheet["G5"].value)
            sheet = workbook["異常類別柏拉圖"]
            self.assertEqual(
                ["排名", "異常類別", "件數", "佔比(%)", "累積佔比(%)"],
                [sheet.cell(row=1, column=col).value for col in range(1, 6)],
            )
            self.assertEqual([1, "物料/來料品質異常", 1, 100.0, 100.0], [
                sheet.cell(row=2, column=col).value for col in range(1, 6)
            ])
            # 表格必須以與頁面圖表相同的區間參數取自同一實作
            mock_pareto.assert_called_once_with("2026-06-01", "2026-06-30")

            self.assertNotIn("異常事件明細", workbook.sheetnames)
            self.assertIn("訪廠", workbook.sheetnames)
            self.assertIn("異常", workbook.sheetnames)

            visit_sheet = workbook["訪廠"]
            self.assertEqual(
                [
                    "日期",
                    "供應商",
                    "問題/摘要",
                    "狀態",
                ],
                [visit_sheet.cell(row=1, column=col).value for col in range(1, 5)],
            )
            self.assertEqual(2, visit_sheet.max_row)
            self.assertEqual(
                ["2026-06-15", "供應商B", "例行訪廠交流", "已完成"],
                [visit_sheet.cell(row=2, column=col).value for col in range(1, 5)],
            )

            anomaly_sheet = workbook["異常"]
            self.assertEqual(
                [
                    "異常單號",
                    "日期",
                    "供應商",
                    "品名",
                    "料號",
                    "階段",
                    "異常類別",
                    "問題/摘要",
                    "確認事項 / 待追蹤",
                    "責任人",
                    "狀態",
                    "品質異常單要求",
                    "改善說明",
                    "結案日期",
                ],
                [anomaly_sheet.cell(row=1, column=col).value for col in range(1, 15)],
            )
            self.assertEqual(4, anomaly_sheet.max_row)
            self.assertEqual(
                [
                    "AN-2026-001",
                    "2026-06-10",
                    "供應商A",
                    "控制板",
                    "PCBA-001",
                    "試產",
                    "物料/來料品質異常",
                    "進料不合格",
                    "追蹤 8D",
                    "王小明",
                    "已結案",
                    "是",
                    "已要求廠商重工",
                    "2026-06-12",
                ],
                [anomaly_sheet.cell(row=2, column=col).value for col in range(1, 15)],
            )
            self.assertEqual("否", anomaly_sheet.cell(row=3, column=12).value)
            self.assertEqual("未設定", anomaly_sheet.cell(row=4, column=12).value)
            self.assertEqual(
                    len(mock_list_events.return_value),
                    (visit_sheet.max_row - 1) + (anomaly_sheet.max_row - 1),
                )

    def test_list_events_by_range_unmocked_db_export(self) -> None:
        base_tmp_dir = Path("scratch")
        base_tmp_dir.mkdir(parents=True, exist_ok=True)
        db_path = base_tmp_dir / f"sqe_excel_export_test_{uuid4().hex}.db"
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        repository.create_schema(conn)

        supplier_id = repository.create_supplier_record(conn, supplier_name="測試光學廠")
        product_id = repository.create_product_record(
            conn,
            product_code="OPT-9988",
            product_name="高階光學鏡頭",
            product_stage="試產",
            supplier_id=supplier_id,
        )
        anomaly_no = repository.create_anomaly(
            conn,
            anomaly_date="2026-06-20",
            supplier_id=supplier_id,
            product_id=product_id,
            problem_desc="鍍膜剝落與刮傷",
            category="製程控制",
            product_name="高階光學鏡頭",
            product_stage="試產",
            pending_items="追蹤 8D 矯正措施",
            responsible_person="陳工程師",
        )

        with patch("database.connection.get_connection", return_value=conn):
            events = event_service.list_events_by_range("2026-06-01", "2026-06-30")
            anomaly_events = [e for e in events if e.get("event_type") == "ANOMALY"]
            self.assertEqual(len(anomaly_events), 1)
            e = anomaly_events[0]
            self.assertEqual(e.get("product_name"), "高階光學鏡頭")
            self.assertEqual(e.get("product_code"), "OPT-9988")
            self.assertEqual(e.get("product_stage"), "試產")
            self.assertEqual(e.get("pending_items"), "追蹤 8D 矯正措施")
            self.assertEqual(e.get("responsible_person"), "陳工程師")

            with TemporaryDirectory() as tmp_dir:
                file_path = os.path.join(tmp_dir, "real_db_events_report.xlsx")
                ok, msg = event_service.export_events_report(
                    file_path=file_path,
                    start_date="2026-06-01",
                    end_date="2026-06-30",
                )
                self.assertTrue(ok)
                self.assertTrue(os.path.exists(file_path))

                wb = load_workbook(file_path)
                self.assertIn("異常", wb.sheetnames)
                sheet = wb["異常"]
                row2 = [sheet.cell(row=2, column=c).value for c in range(1, 15)]
                self.assertEqual(row2[0], anomaly_no)
                self.assertEqual(row2[1], "2026-06-20")
                self.assertEqual(row2[2], "測試光學廠")
                self.assertEqual(row2[3], "高階光學鏡頭")
                self.assertEqual(row2[4], "OPT-9988")
                self.assertEqual(row2[5], "試產")
                self.assertEqual(row2[6], "製程控制")
                self.assertEqual(row2[7], "鍍膜剝落與刮傷")
                self.assertEqual(row2[8], "追蹤 8D 矯正措施")
                self.assertEqual(row2[9], "陳工程師")

        conn.close()
        if db_path.exists():
            db_path.unlink()
